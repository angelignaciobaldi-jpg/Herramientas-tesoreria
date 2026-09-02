"""Deriva los artefactos del módulo de Saldos a partir del formato de tesorería.

Este script NO corre en la app. Se ejecuta a mano cuando cambia el formato:

    python scripts/derivar_plantilla_saldos.py "descargas/ReporteCuentas/formatoCompleto.xlsx"

Produce dos archivos en `core/datos/`:

  saldos_base.xlsx  El mismo libro, con los datos vaciados. Conserva las 24
                    pestañas, encabezados, fórmulas, estilos, celdas combinadas,
                    tablas, la hoja oculta y la configuración de impresión. Es el
                    molde que `core/saldos_export.py` rellena.

  saldos_mapa.json  El mapeo. Dice, para cada cuenta, EN QUÉ FILA de qué pestaña
                    va, y qué columna de esa fila lleva cada dato.

## Por qué existe

En el formato manual cada saldo de la hoja SALDOS es una referencia a celda fija
(`=HSBC!C2`). La relación cuenta-empresa vive en la POSICIÓN de la fila: si el
portal cambia el orden de la descarga, el reporte sale mal sin avisar. De hecho ya
pasó: cuatro renglones de HSBC del formato apuntan hoy a la cuenta equivocada.

La solución no es tirar las fórmulas, sino quitarles el supuesto: si nosotros
colocamos cada cuenta en su fila canónica, `=HSBC!C2` vuelve a ser correcto POR
CONSTRUCCIÓN. Para eso hace falta saber qué cuenta va en cada fila, y eso es lo
que este script extrae: sigue cada fórmula de SALDOS hasta su hoja de banco y de
ahí saca el número de cuenta y el titular reales.

## Lo que no se puede automatizar

Cada portal pega distinto, así que las REGIONES de cada pestaña (dónde empieza el
bloque de datos y qué columna es cuál) van a mano en la tabla `REGIONES`. Es la
parte irreducible: no hay forma de adivinar que en BAJÍO el número de cuenta vive
en la fila de ABAJO, o que 'BX+ SCO' tiene tres bancos en una sola hoja.
"""

from __future__ import annotations

import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import column_index_from_string

_RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _RAIZ not in sys.path:
    sys.path.insert(0, _RAIZ)

DESTINO = os.path.join(_RAIZ, "core", "datos")


# ---------------------------------------------------------------------------
# Regiones: dónde vive cada dato en cada pestaña de descarga
# ---------------------------------------------------------------------------

@dataclass
class Region:
    """Un bloque de datos pegado dentro de una pestaña.

    `fila_ini`/`fila_fin` son inclusivos y NO incluyen el encabezado. `cols` mapea
    rol -> letra de columna; los roles que una región no tiene simplemente no
    aparecen. `moneda_fija` se usa cuando el bloque entero es de una divisa y el
    portal no la repite por fila (Banamex separa MXN y USD en bloques distintos).

    `cuenta_abajo` es el caso de BAJÍO: el número no está en la fila del saldo
    sino en la siguiente, con la forma "Cuenta: 16084470201Conecta BanBajio".
    """

    fila_ini: int
    fila_fin: int
    cols: dict
    moneda_fija: str = None
    cuenta_abajo: bool = False
    paso: int = 1          # BAJÍO avanza de 2 en 2 (dato, cuenta, dato, cuenta...)
    nota: str = ""

    def filas(self):
        return range(self.fila_ini, self.fila_fin + 1, self.paso)

    def columnas(self):
        """Rango de columnas que ocupa, para poder vaciarla."""
        letras = list(self.cols.values())
        if not letras:
            return None
        idx = [column_index_from_string(x) for x in letras]
        return min(idx), max(idx)


REGIONES = {
    "AFIRME": [
        Region(3, 3, {"descripcion": "A", "cuenta": "B", "clabe": "C",
                      "moneda": "D", "titular": "E", "saldo": "F",
                      "saldo_total": "G"}),
    ],
    # Banamex parte MXN y USD en dos bloques; la cuenta va sin la sucursal, que
    # está en su propia columna. El lector concatena las dos (`3947680454`), por
    # eso el casado tiene que probar también `extra["cuenta_corta"]`.
    "BANAMEX": [
        Region(3, 16, {"titular": "B", "sucursal": "C", "cuenta": "D",
                       "saldo": "E"}, moneda_fija="MXN"),
        Region(19, 20, {"titular": "B", "sucursal": "C", "cuenta": "D",
                        "saldo": "E"}, moneda_fija="USD"),
    ],
    # BAJÍO: el portal exporta nombre+saldo en una fila y el número en la de
    # abajo. Las columnas E-G e I-K son copias de otras descargas que el formato
    # no referencia; se vacían y no se vuelven a escribir.
    "BAJIO": [
        Region(2, 20, {"titular": "A", "retenido": "B", "saldo": "C"},
               moneda_fija="MXN", cuenta_abajo=True, paso=2,
               nota="el número va en la fila siguiente, tras 'Cuenta: '"),
    ],
    "BANCOPPEL": [
        Region(1, 1, {"cuenta": "C", "clabe": "F", "saldo": "I"},
               moneda_fija="MXN"),
    ],
    "BANREGIO": [
        Region(2, 41, {"cuenta": "A", "titular": "B", "saldo": "C"},
               moneda_fija="MXN"),
        Region(45, 48, {"cuenta": "A", "titular": "B", "saldo": "C"},
               moneda_fija="MXN", nota="bloque de cuentas nuevas"),
    ],
    "BANORTE": [
        Region(2, 25, {"cuenta": "A", "titular": "B", "moneda": "C",
                       "clabe": "D", "saldo_actual": "E", "saldo": "F"}),
    ],
    # BBVA se descarga en varios archivos y cada uno se pega en su propio bloque.
    # El primero trae cuentas de 9 dígitos y el saldo en F; los demás, cuentas de
    # 18 dígitos y el saldo en E.
    "BANCOMER": [
        Region(2, 21, {"cuenta": "A", "titular": "B", "moneda": "C",
                       "saldo_actual": "E", "saldo": "F"}),
        Region(28, 45, {"cuenta": "A", "titular": "B", "moneda": "C",
                        "saldo_actual": "D", "saldo": "E"}),
        Region(53, 70, {"cuenta": "A", "titular": "B", "moneda": "C",
                        "saldo_actual": "D", "saldo": "E"}),
        Region(74, 91, {"cuenta": "A", "titular": "B", "moneda": "C",
                        "saldo_actual": "D", "saldo": "E"}),
        Region(94, 103, {"cuenta": "A", "titular": "B", "moneda": "C",
                         "saldo_actual": "D", "saldo": "E"}),
    ],
    # HSBC: filas 2-21 del portal, más 23-24 que la usuaria captura a mano (no
    # traen número de cuenta; se resuelven por el puente).
    "HSBC": [
        Region(2, 21, {"cuenta": "A", "titular": "B", "saldo": "C",
                       "saldo_contable": "F"}),
        Region(23, 24, {"titular": "B", "saldo": "C"},
               nota="capturadas a mano, sin número de cuenta"),
    ],
    "INBURSA": [
        Region(5, 6, {"cuenta": "A", "titular": "B", "saldo_inicial": "C",
                      "retenido": "D", "buen_cobro": "E", "saldo": "F"}),
    ],
    "INTERCAM": [
        Region(2, 2, {"cuenta": "A", "moneda": "B", "titular": "C",
                      "saldo": "D", "sobregiro": "E"},
               nota="el portal enmascara la cuenta: '***-***94-001-1'"),
    ],
    # Monex identifica por alias ('ABACOM 0875'), no por número. Además el
    # formato desglosa MN / DLLS / FORWARDS y el PDF del portal trae un solo
    # total, así que hay más filas que archivos.
    "MONEX": [
        Region(3, 8, {"alias": "A", "efectivo": "B", "otras_divisas": "C",
                      "derivados": "D", "inversion": "E", "saldo": "F"}),
    ],
    "MULTIVA": [
        Region(2, 2, {"titular": "A", "cuenta": "B", "alias": "C",
                      "moneda": "D", "saldo": "E"}),
    ],
    # Sabadell alterna fila de etiqueta ('Cuenta Expansión MXN') y fila de datos.
    "SABADELL": [
        Region(9, 9, {"cuenta": "A", "titular": "B", "estado": "C",
                      "moneda": "D", "saldo": "E"}),
        Region(11, 11, {"cuenta": "A", "titular": "B", "estado": "C",
                        "moneda": "D", "saldo": "E"}),
    ],
    # El bloque de abajo (filas 18-28) es una SEGUNDA descarga del mismo día, con
    # las mismas cuentas y otro corte de hora. Ninguna fórmula lo lee, y como solo
    # tenemos un corte, llenarlo con los mismos números fingiría un dato que no
    # tenemos. Se vacía y no se vuelve a escribir, igual que los bloques de más
    # de BAJÍO.
    "SANTANDER": [
        Region(2, 15, {"cuenta": "B", "titular": "C", "hora": "D",
                       "saldo": "E"}, moneda_fija="MXN"),
    ],
    # Tres bancos distintos conviven en una hoja. SALDOS solo referencia las
    # secciones de Scotiabank (G) y BX+ (E); la de Bajío de abajo no se usa.
    "BX+ SCO": [
        Region(3, 5, {"producto": "A", "pais": "B", "plaza": "C", "cuenta": "D",
                      "titular": "E", "moneda": "F", "saldo": "G"},
               nota="Scotiabank"),
        Region(8, 9, {"producto": "B", "titular": "C", "moneda": "D",
                      "saldo": "E"}, nota="Ve por Mas (BX+)"),
    ],
}

# Rangos que hay que vaciar aunque no sean regiones de datos: sobras de pegados
# anteriores y totales capturados a mano. Ninguna fórmula del formato los usa, y
# si se quedan muestran cifras viejas junto a las nuevas.
#
#   BANAMEX    el portal exporta dos tablas más (G-N y G-K) que nadie referencia
#   BAJIO      las columnas E-G e I-K son copias de otras descargas
#   BX+ SCO    la sección de Bajío de abajo, y una URL que quedó en D3
#   los demás  totales tecleados a mano al pie de cada bloque
RESIDUOS = {
    "BANAMEX": [(2, 21, "G", "N"), (27, 39, "G", "K")],
    "BAJIO": [(2, 17, "E", "G"), (2, 17, "I", "K")],
    "BX+ SCO": [(3, 3, "D", "D"), (12, 17, "A", "G")],
    "BANCOMER": [(22, 23, "E", "F")],
    "SANTANDER": [(15, 15, "G", "G"), (18, 28, "B", "G")],
    "BANREGIO": [(23, 24, "F", "F"), (44, 44, "C", "C"), (45, 46, "D", "D")],
    "AFIRME": [(4, 4, "F", "G")],
    "BANORTE": [(26, 27, "A", "F")],
}

# Pestañas que llevan insumos de flujo, no descargas bancarias. Las llena
# `core/saldos_insumos.py`.
#
# Cuatro de ellas son ledgers de vencimientos y la hoja SALDOS las consulta
# siempre igual: `SUMIF(<fecha>, día, <importe>)`. Solo hace falta saber qué
# columna es cuál, y eso es lo que declara `cols`.
#
# CRÉDITOS es distinta: no es una descarga sino la hoja que tesorería mantiene a
# mano, y SALDOS la lee por posición (`=+CRÉDITOS!F9`). Se copia tal cual, en las
# mismas coordenadas, para que esas referencias sigan siendo válidas. Es
# aceptable aquí y no en las pestañas de banco porque quien fija esas filas es
# una persona, no el orden en que un portal decidió exportar.
# Hasta dónde puede crecer un ledger. Los `SUMIF` de SALDOS barren la columna
# entera (`SUMIF(PEMEX!D:D, ...)`), así que pasarse del último renglón que traía
# el formato no rompe nada; ajustar el tope al tamaño de la descarga de ese día
# sí habría reventado con la siguiente, que trae unas filas más.
_TOPE_LEDGER = 100000

LEDGERS = {
    "MGC": {"fila_ini": 2, "fila_fin": _TOPE_LEDGER, "col_ini": "A", "col_fin": "L",
            "cols": {"referencia": "A", "importe": "I", "fecha": "J"}},
    "PEMEX": {"fila_ini": 2, "fila_fin": _TOPE_LEDGER, "col_ini": "A", "col_fin": "M",
              "cols": {"referencia": "A", "fecha": "D", "importe": "H"}},
    "TESORO": {"fila_ini": 2, "fila_fin": _TOPE_LEDGER, "col_ini": "A", "col_fin": "J",
               "cols": {"referencia": "A", "importe": "B", "fecha": "E"}},
    "NOMINA": {"fila_ini": 2, "fila_fin": _TOPE_LEDGER, "col_ini": "A", "col_fin": "J",
               "cols": {"referencia": "A", "fecha": "C", "importe": "D"}},
    "IMPUESTOS": {"fila_ini": 2, "fila_fin": 24, "col_ini": "B", "col_fin": "H",
                  "cols": {}, "nota": "el bloque existe pero SALDOS no lo lee"},
    # CRÉDITOS no es una descarga: es una hoja que tesorería mantiene a mano, y
    # casi todo lo que hay en ella es ESQUELETO, no dato. Vaciarla por rectángulos
    # dejaba la pestaña desnuda —se perdían los títulos de sección ('ABASTECEDORA',
    # 'PETROPLAZAS'…), los nombres de banco de la columna A, las cuentas de la B,
    # los 'Total:' y las 208 etiquetas empresa/acreedor de la matriz J/K—, y con
    # ella el formato que se ofrece a descargar: quien lo abría se encontraba una
    # cuadrícula en blanco sin saber qué va en cada renglón.
    #
    # La hoja no tiene NI UNA fórmula (comprobado sobre el formato): todo es texto
    # o número tecleado. Eso permite una regla exacta —y por eso el modo
    # 'solo_numeros'—: en las columnas de importes se borra lo que no sea texto, y
    # las columnas de identidad (A, B, J, K) no se tocan nunca. Se comprobó que en
    # C-F filas 8-139 el único texto son los títulos y los 'Total:', y que en L-V
    # lo único no numérico son las fechas de la semana (M3:V3), que sí son dato.
    "CREDITOS": {"fila_ini": 2, "fila_fin": 256, "col_ini": "A", "col_fin": "V",
                 "cols": {}, "modo": "copia",
                 "rangos": [[2, 139, "C", "F", "solo_numeros"],
                            [2, 118, "L", "V", "solo_numeros"]]},
}

# Totales de la cabecera de SALDOS. Las filas 3-4 son FÓRMULAS y valen para hoy;
# las 5-6 son VALORES que tesorería pega a mano con los totales del día hábil
# anterior. El mapa guarda a qué celda de las filas 5-6 le toca cada total de las
# filas 3-4, para que la app pueda hacer ese copiado sola.
#
# Ojo con la asimetría: en pesos el desglose 'Combustibles' va en S y en dólares
# en R. No es un error de captura, así viene el formato.
ESPEJO_TOTALES = {
    "Q3": "Q5",   # MX  · total
    "S3": "S5",   # MX  · Combustibles
    "U3": "U5",   # MX  · Resto
    "Q4": "Q6",   # DLS · total
    "R4": "R6",   # DLS · Combustibles
    "U4": "U6",   # DLS · Resto
}
# Dónde va la fecha y la hora de la corrida anterior.
CELDA_FECHA_ANTERIOR = "L5"
CELDA_HORA_ANTERIOR = "L6"

# Bandas de la hoja SALDOS: (nombre, banco, cuenta, titulo, saldo, moneda).
# En la banda derecha el título y el banco comparten la columna Q; se distinguen
# porque el título va en una celda combinada Q:U.
BANDAS = [
    ("A-I", "A", "B", "F", "I", "H"),
    ("K-O", "K", "L", "M", "O", "N"),
    ("Q-U", "Q", "R", "Q", "U", "T"),
]

_MARCAS_MONEDA = {"DLS", "DLLS", "USD", "TDE", "CLN", "MZT", "LM"}

# Cuentas que el formato no puede dar por sí solo, con su número real:
#
#   MONEX     identifica por alias ('ABACOM 0875'); el número sale del PDF.
#   BX+ SCO   la celda de Scotiabank trae el número mutilado (le falta el 1
#             inicial: '1700512613' en vez de '11700512613').
#   HSBC      las filas 23-24 se capturan a mano y no traen número.
#
# Las demás filas sin número (Monex DLLS/FORWARDS, BX+, Santander 14-15, Banamex
# 16, Intercam) se resuelven por (banco, últimos 4) al momento de casar.
PUENTE = {
    ("MONEX", 3): "2750875",
    ("MONEX", 4): "2793487",
    ("MONEX", 5): "2848000",
    ("MONEX", 8): "2930014",
    ("BX+ SCO", 3): "11700512613",
    ("BX+ SCO", 5): "25605315790",
    ("HSBC", 23): "4056511132",
    ("HSBC", 24): "4060386885",
}

# Debajo de este número de dígitos, lo que se sacó de la celda no es un número de
# cuenta utilizable: es una máscara del portal ('***-***94-001-1') o basura.
_MIN_DIGITOS_CUENTA = 6

# Roles cuyo valor CAMBIA cada día: son los que trae el archivo del portal. Todo
# lo demás —número, titular, sucursal, CLABE, plaza, tipo de producto— es la
# IDENTIDAD de la cuenta, no varía, y se conserva del formato.
#
# La distinción importa: al vaciar el libro base se iba también la identidad, así
# que una cuenta que el portal no reportara ese día desaparecía por completo de su
# pestaña en vez de aparecer con el saldo en blanco. Un renglón vacío se ve; una
# cuenta que no está, no.
ROLES_SALDO = frozenset({
    "saldo", "saldo_actual", "saldo_total", "saldo_inicial", "saldo_contable",
    "retenido", "buen_cobro", "sobregiro",
    "efectivo", "otras_divisas", "derivados", "inversion",
})


def _sin_acentos(texto):
    plano = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in plano if not unicodedata.combining(c))


def clave_hoja(nombre):
    """Nombre de hoja normalizado: 'CRÉDITOS' y 'CREDITOS' son lo mismo."""
    return _sin_acentos(nombre).upper().strip()


def digitos(valor):
    return re.sub(r"\D", "", str(valor or ""))


def normalizar_cuenta(valor):
    """Número comparable: solo dígitos y sin ceros a la izquierda.

    Los portales rellenan con ceros de forma inconsistente — BBVA exporta
    `000000000110311944` en un archivo y `110311944` en otro."""
    d = digitos(valor)
    return d.lstrip("0") or d


# ---------------------------------------------------------------------------
# Lectura del formato
# ---------------------------------------------------------------------------

_REF = re.compile(r"^=\+?'?([^'!]+?)'?!\$?([A-Z]{1,2})\$?(\d+)$")


def _resolver_ref(formula):
    """Parte '=HSBC!C2' en ('HSBC', 'C', 2). Devuelve None si no es esa forma."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    m = _REF.match(formula.strip())
    if not m:
        return None
    return m.group(1).strip(), m.group(2), int(m.group(3))


def _celda(hoja, col, fila):
    return hoja["{}{}".format(col, fila)].value


def leer_cuentas_por_fila(libro):
    """Por pestaña de descarga: {fila -> datos de esa fila}, según sus regiones.

    Es el inventario de qué cuenta ocupa hoy cada fila del formato. De aquí sale
    tanto el orden canónico como los números completos que necesita el casado."""
    inventario = {}
    for nombre, regiones in REGIONES.items():
        hoja = libro[nombre]
        filas = {}
        for i, region in enumerate(regiones):
            for fila in region.filas():
                datos = {}
                for rol, col in region.cols.items():
                    valor = _celda(hoja, col, fila)
                    if valor is not None:
                        datos[rol] = valor
                if not datos:
                    continue
                if region.cuenta_abajo:
                    crudo = _celda(hoja, region.cols["titular"], fila + 1)
                    datos["cuenta"] = digitos(
                        str(crudo or "").split("Cuenta:")[-1][:20])
                # Identidad: lo que NO es saldo se conserva tal cual del formato.
                datos["estaticos"] = {
                    rol: valor for rol, valor in datos.items()
                    if rol not in ROLES_SALDO and valor is not None
                }
                if region.cuenta_abajo:
                    # En BAJÍO el número vive en la fila de abajo ('Cuenta:
                    # 16084470201Conecta BanBajio'). Esa celda también es
                    # identidad y hay que reponerla, o la pestaña sale coja.
                    crudo = _celda(hoja, region.cols["titular"], fila + 1)
                    if crudo is not None:
                        datos["estaticos"]["_fila_cuenta_abajo"] = crudo
                if region.moneda_fija and "moneda" not in datos:
                    datos["moneda"] = region.moneda_fija
                    datos["estaticos"]["moneda"] = region.moneda_fija
                # Una máscara del portal ('***-***94-001-1') deja dígitos sueltos
                # que parecerían una cuenta y casarían con cualquier cosa.
                if len(digitos(datos.get("cuenta"))) < _MIN_DIGITOS_CUENTA:
                    datos.pop("cuenta", None)
                puente = PUENTE.get((nombre, fila))
                if puente:
                    datos["cuenta"] = puente
                    datos["por_puente"] = True
                datos["region"] = i
                filas[fila] = datos
        inventario[nombre] = filas
    return inventario


def leer_totales_cabecera(libro, renglones):
    """Qué celdas de saldo suma, AL FINAL, cada total de la cabecera.

    Las fórmulas del formato van en dos pisos: `Q3` suma totales de bloque
    (`I30`), y cada total de bloque suma sus renglones (`=SUM(I10:I29)`). Aquí se
    aplanan hasta las celdas hoja —las 209 que llevan un saldo— resolviendo los
    `SUM` intermedios contra el propio formato.

    Se hace UNA vez, aquí, y no al generar: así la app solo tiene que sumar los
    saldos que colocó, sin evaluar fórmulas de Excel en tiempo de ejecución.

    El desglose Combustibles/Resto sale de las fórmulas y no se reconstruye: es el
    criterio que definió tesorería, aunque ya nadie recuerde por qué esas empresas
    y no otras."""
    hoja = libro["SALDOS"]
    hojas_banco = set(REGIONES)
    suelta = re.compile(r"\$?([A-Z]{1,2})\$?(\d+)(?!\d)")
    rango = re.compile(r"\$?([A-Z]{1,2})\$?(\d+):\$?([A-Z]{1,2})\$?(\d+)")

    def es_hoja(celda):
        """Una celda hoja lee directo de una pestaña de banco."""
        v = hoja[celda].value
        ref = _resolver_ref(v) if isinstance(v, str) else None
        return ref is not None and clave_hoja(ref[0]) in hojas_banco

    def expandir(celda, visto):
        if celda in visto:      # el formato tiene sumas que se repiten
            return []
        visto = visto | {celda}
        if es_hoja(celda):
            return [celda]
        formula = hoja[celda].value
        if not isinstance(formula, str) or not formula.startswith("="):
            return []
        hijas = []
        cuerpo = formula[1:]
        for c1, f1, c2, f2 in rango.findall(cuerpo):
            for col in range(column_index_from_string(c1),
                             column_index_from_string(c2) + 1):
                letra = openpyxl.utils.get_column_letter(col)
                for f in range(int(f1), int(f2) + 1):
                    hijas.append("{}{}".format(letra, f))
            cuerpo = cuerpo.replace(
                "{}{}:{}{}".format(c1, f1, c2, f2), " ", 1)
        hijas += ["{}{}".format(c, f) for c, f in suelta.findall(cuerpo)]
        salida = []
        for h in hijas:
            salida.extend(expandir(h, visto))
        return salida

    celdas_saldo = {"{}{}".format(
        {b[0]: b[4] for b in BANDAS}[r["banda"]], r["fila"]) for r in renglones}

    totales, avisos = {}, []
    for celda in ESPEJO_TOTALES:
        formula = hoja[celda].value
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        hojas_ = expandir(celda, frozenset())
        # Solo se conservan las que de verdad llevan un saldo del reporte: así un
        # dedazo del formato que apunte a una celda vacía no entra al total.
        buenas = [c for c in hojas_ if c in celdas_saldo]
        if len(buenas) != len(hojas_):
            avisos.append((celda, sorted(set(hojas_) - celdas_saldo)))
        totales[celda] = buenas
    return totales, avisos


def leer_renglones(libro):
    """Los renglones de la hoja SALDOS: uno por cada celda de saldo con fórmula.

    El título del bloque se arrastra hacia abajo: en el formato aparece una vez,
    arriba del grupo, y aplica hasta el siguiente título."""
    hoja = libro["SALDOS"]
    # Los títulos de la banda derecha van en una celda combinada que arranca en Q
    # y llega hasta U o hasta V, según el bloque. Nos basta con dónde empieza.
    titulos_combinados = {r.min_row for r in hoja.merged_cells.ranges
                          if r.min_col == column_index_from_string("Q")
                          and r.max_col > r.min_col}
    renglones = []
    for banda, c_banco, c_cuenta, c_titulo, c_saldo, c_moneda in BANDAS:
        bloque = None
        for fila in range(1, hoja.max_row + 1):
            titulo = _celda(hoja, c_titulo, fila)
            saldo = _celda(hoja, c_saldo, fila)
            tiene_saldo = isinstance(saldo, str) and saldo.startswith("=")
            # En la banda derecha el título va en una celda combinada Q:U; en las
            # otras dos, en una columna que no comparte con el banco.
            es_titulo = (isinstance(titulo, str) and titulo.strip()
                         and not titulo.startswith("="))
            if banda == "Q-U":
                es_titulo = es_titulo and fila in titulos_combinados
            else:
                es_titulo = es_titulo and not tiene_saldo
            if es_titulo:
                bloque = titulo.strip()
                continue
            if not tiene_saldo:
                continue
            ref = _resolver_ref(saldo)
            if ref is None or clave_hoja(ref[0]) not in REGIONES:
                continue
            marca = _celda(hoja, c_moneda, fila)
            marca = str(marca).strip().upper() if marca else ""
            renglones.append({
                "banda": banda,
                "fila": fila,
                "bloque": bloque,
                "banco": str(_celda(hoja, c_banco, fila) or "").strip(),
                "cta4": str(_celda(hoja, c_cuenta, fila) or "").strip(),
                "marca": marca if marca in _MARCAS_MONEDA else "",
                "ref_hoja": ref[0],
                "ref_col": ref[1],
                "ref_fila": ref[2],
            })
    return renglones


# ---------------------------------------------------------------------------
# Desfases: renglones cuya referencia apunta a la cuenta equivocada
# ---------------------------------------------------------------------------

def detectar_desfases(renglones, inventario):
    """Encuentra referencias que apuntan a una fila que NO es la de esa cuenta.

    Es el defecto que motivó todo esto: cuando la descarga de un portal cambia de
    orden, `=HSBC!C18` sigue apuntando a la fila 18 aunque ahí ya viva otra
    cuenta. El formato actual tiene cinco renglones de HSBC así.

    Se detecta con el dato que el propio formato ya trae: la columna de cuenta del
    renglón (los últimos 4 dígitos). Si en esa pestaña hay UNA sola fila cuya
    cuenta termina en esos 4 dígitos, y no es la referenciada, la referencia está
    mal. Se exige unicidad porque bancos como Banregio tienen decenas de cuentas
    que comparten cola.
    """
    desfases = []
    for r in renglones:
        cola = digitos(r["cta4"])[-4:]
        if len(cola) < 4:
            continue
        filas = inventario.get(clave_hoja(r["ref_hoja"]), {})
        candidatas = [f for f, d in filas.items()
                      if digitos(d.get("cuenta"))[-4:] == cola]
        if len(candidatas) != 1:
            continue
        correcta = candidatas[0]
        if correcta == r["ref_fila"]:
            continue
        # Solo es desfase si la fila referenciada tiene OTRA cuenta. Si está
        # vacía, puede ser una fila que el portal no trajo ese día.
        actual = filas.get(r["ref_fila"], {})
        if not digitos(actual.get("cuenta")):
            continue
        desfases.append({
            "banda": r["banda"], "fila": r["fila"], "bloque": r["bloque"],
            "banco": r["banco"], "cta4": r["cta4"],
            "ref_mala": "{}!{}{}".format(r["ref_hoja"], r["ref_col"],
                                         r["ref_fila"]),
            "ref_buena": "{}!{}{}".format(r["ref_hoja"], r["ref_col"], correcta),
            "cuenta_que_leia": str(actual.get("cuenta")),
            "cuenta_correcta": str(filas[correcta].get("cuenta")),
            "ref_fila_buena": correcta,
        })
    return desfases


# ---------------------------------------------------------------------------
# Ensamblado del mapa
# ---------------------------------------------------------------------------

def armar_mapa(renglones, inventario, desfases, totales_cabecera=None):
    """Junta renglones e inventario en la estructura que consume la app."""
    totales_cabecera = totales_cabecera or {}
    por_celda = {(d["banda"], d["fila"]): d for d in desfases}

    hojas = {}
    for nombre, regiones in REGIONES.items():
        filas_json = {}
        for fila, datos in sorted(inventario[nombre].items()):
            entrada = {"region": datos["region"]}
            if datos.get("cuenta"):
                entrada["cuenta"] = str(datos["cuenta"])
                entrada["cuenta_norm"] = normalizar_cuenta(datos["cuenta"])
            for rol in ("titular", "alias", "moneda"):
                if datos.get(rol) is not None:
                    entrada[rol] = str(datos[rol]).strip()
            if datos.get("por_puente"):
                entrada["por_puente"] = True
            entrada["estaticos"] = {
                rol: (valor.isoformat() if hasattr(valor, "isoformat") else valor)
                for rol, valor in (datos.get("estaticos") or {}).items()
            }
            filas_json[str(fila)] = entrada
        hojas[nombre] = {
            "regiones": [
                {"fila_ini": g.fila_ini, "fila_fin": g.fila_fin,
                 "paso": g.paso, "cols": g.cols,
                 "moneda_fija": g.moneda_fija, "cuenta_abajo": g.cuenta_abajo,
                 "nota": g.nota}
                for g in regiones
            ],
            "filas": filas_json,
        }

    renglones_json = []
    for r in renglones:
        hoja = clave_hoja(r["ref_hoja"])
        fila_ref = r["ref_fila"]
        arreglo = por_celda.get((r["banda"], r["fila"]))
        if arreglo:
            fila_ref = arreglo["ref_fila_buena"]
        datos = inventario.get(hoja, {}).get(fila_ref, {})
        renglones_json.append({
            "banda": r["banda"],
            "fila": r["fila"],
            "bloque": r["bloque"],
            "banco": r["banco"],
            "cta4": r["cta4"],
            "marca": r["marca"],
            "hoja": hoja,
            "hoja_fila": fila_ref,
            "hoja_col": r["ref_col"],
            "cuenta": str(datos.get("cuenta") or "") or None,
            "titular": str(datos.get("titular") or datos.get("alias") or "")
                       or None,
            "reparado": bool(arreglo),
        })

    return {
        "version": 2,
        "origen": "derivado de la hoja SALDOS del formato de tesorería",
        "bandas": [
            {"nombre": b[0], "banco": b[1], "cuenta": b[2], "titulo": b[3],
             "saldo": b[4], "moneda": b[5]}
            for b in BANDAS
        ],
        "ledgers": LEDGERS,
        "hojas": hojas,
        "totales_cabecera": totales_cabecera,
        "espejo_totales": ESPEJO_TOTALES,
        "celda_fecha_anterior": CELDA_FECHA_ANTERIOR,
        "celda_hora_anterior": CELDA_HORA_ANTERIOR,
        "renglones": renglones_json,
        "desfases_corregidos": desfases,
    }


# ---------------------------------------------------------------------------
# Vaciado del libro base
# ---------------------------------------------------------------------------

def _limpiar(hoja, fila_ini, fila_fin, col_ini, col_fin, modo=""):
    """Borra el CONTENIDO de un rango sin tocar formato ni celdas combinadas.

    Con `modo="solo_numeros"` respeta el texto: sirve para las hojas donde los
    rótulos comparten columna con los importes y no se pueden separar por
    rectángulos (ver CRÉDITOS en LEDGERS)."""
    fila_fin = min(fila_fin, hoja.max_row)
    for fila in range(fila_ini, fila_fin + 1):
        for col in range(col_ini, col_fin + 1):
            celda = hoja.cell(fila, col)
            if celda.value is None:
                continue
            if modo == "solo_numeros" and isinstance(celda.value, str):
                continue
            celda.value = None


def vaciar(libro):
    """Deja el libro como molde: sin datos, con toda su estructura.

    Se vacían las 15 pestañas de descarga y los 6 ledgers. NO se toca SALDOS ni
    SALDOS HORIZOTAL: ahí todo es fórmula, y es justo lo que queremos conservar.
    """
    for nombre, regiones in REGIONES.items():
        hoja = libro[nombre]
        for region in regiones:
            rango = region.columnas()
            if rango is None:
                continue
            # BAJÍO guarda el número en la fila de abajo: hay que barrer también
            # las filas intermedias, no solo las del paso.
            _limpiar(hoja, region.fila_ini,
                     region.fila_fin + (1 if region.cuenta_abajo else 0),
                     rango[0], rango[1])
        for fi, ff, ci, cf in RESIDUOS.get(nombre, ()):
            _limpiar(hoja, fi, ff, column_index_from_string(ci),
                     column_index_from_string(cf))

    # Hoja1 es una lista auxiliar oculta y NADA la referencia (se verificó sobre
    # el formato). Sus referencias rotas son ruido heredado: se blanquean.
    hoja1 = libro["Hoja1"]
    for fila in range(1, hoja1.max_row + 1):
        for col in range(1, hoja1.max_column + 1):
            celda = hoja1.cell(fila, col)
            if isinstance(celda.value, str) and "#REF!" in celda.value:
                celda.value = None

    for nombre, info in LEDGERS.items():
        hoja = _hoja(libro, nombre)
        # CRÉDITOS solo se vacía en los rangos que de verdad son datos: sus filas
        # 1-6 llevan encabezados y parámetros de control que SALDOS consulta.
        rangos = info.get("rangos") or [[info["fila_ini"], info["fila_fin"],
                                         info["col_ini"], info["col_fin"]]]
        for rango in rangos:
            fi, ff, ci, cf = rango[:4]
            modo = rango[4] if len(rango) > 4 else ""
            _limpiar(hoja, fi, ff, column_index_from_string(ci),
                     column_index_from_string(cf), modo)


def _hoja(libro, clave):
    """Busca una hoja por su nombre normalizado ('CREDITOS' -> 'CRÉDITOS')."""
    for nombre in libro.sheetnames:
        if clave_hoja(nombre) == clave:
            return libro[nombre]
    raise KeyError("no existe la hoja {!r}".format(clave))


def reparar_referencias(libro, desfases):
    """Reescribe en el base las fórmulas que apuntaban a la cuenta equivocada."""
    hoja = libro["SALDOS"]
    bandas = {b[0]: b for b in BANDAS}
    for d in desfases:
        col = bandas[d["banda"]][4]
        celda = hoja["{}{}".format(col, d["fila"])]
        celda.value = "={}".format(d["ref_buena"])


# ---------------------------------------------------------------------------

def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    origen = argv[1]
    if not os.path.exists(origen):
        print("no existe: {}".format(origen))
        return 2

    print("leyendo {}".format(origen))
    con_valores = openpyxl.load_workbook(origen, data_only=True)
    con_formulas = openpyxl.load_workbook(origen, data_only=False)

    inventario = leer_cuentas_por_fila(con_valores)
    renglones = leer_renglones(con_formulas)
    desfases = detectar_desfases(renglones, inventario)

    print("\n{} renglones en SALDOS, {} bloques".format(
        len(renglones), len({r["bloque"] for r in renglones})))
    sin_cuenta = [r for r in renglones
                  if not inventario.get(clave_hoja(r["ref_hoja"]), {})
                  .get(r["ref_fila"], {}).get("cuenta")]
    print("{} renglones con cuenta resuelta, {} sin ella".format(
        len(renglones) - len(sin_cuenta), len(sin_cuenta)))

    if desfases:
        print("\n{} referencia(s) apuntaban a la cuenta equivocada:".format(
            len(desfases)))
        for d in desfases:
            print("  {} {:<26} {:<10} {} -> {}  (leia {}, debe leer {})".format(
                d["banda"], (d["bloque"] or "")[:26], d["cta4"],
                d["ref_mala"], d["ref_buena"],
                d["cuenta_que_leia"], d["cuenta_correcta"]))

    if sin_cuenta:
        print("\n{} renglon(es) sin numero de cuenta; se casaran por "
              "(banco, ultimos 4):".format(len(sin_cuenta)))
        for r in sin_cuenta:
            print("  {} f{:<4} {:<26} {:<12} {:<9} -> {}!{}{}".format(
                r["banda"], r["fila"], (r["bloque"] or "")[:26],
                r["banco"][:12], r["cta4"], r["ref_hoja"], r["ref_col"],
                r["ref_fila"]))

    totales_cabecera, avisos_totales = leer_totales_cabecera(
        con_formulas, renglones)
    if avisos_totales:
        print("\ntotales de cabecera que apuntan a celdas SIN saldo "
              "(se excluyen de la suma):")
        for celda, sobras in avisos_totales:
            print("  {}: {}".format(celda, ", ".join(sobras)))
    print("\ntotales de cabecera: " + ", ".join(
        "{} suma {} renglones".format(k, len(v))
        for k, v in sorted(totales_cabecera.items())))

    mapa = armar_mapa(renglones, inventario, desfases, totales_cabecera)

    os.makedirs(DESTINO, exist_ok=True)
    ruta_json = os.path.join(DESTINO, "saldos_mapa.json")
    with open(ruta_json, "w", encoding="utf-8") as f:
        json.dump(mapa, f, ensure_ascii=False, indent=1)
    print("\nescrito {}".format(ruta_json))

    reparar_referencias(con_formulas, desfases)
    cambios, dudas = reparar_horizontal(con_formulas, inventario, renglones, desfases)
    if cambios:
        print("\n{} referencia(s) corregidas en SALDOS HORIZOTAL:".format(
            len(cambios)))
        for c in cambios:
            print("  {:<6} {:<4} {:<34} {} -> {}".format(
                c["celda"], c["divisa"], c["empresa"][:34], c["antes"][:34],
                c["despues"][:34]))
    if dudas:
        print("\n{} referencia(s) de SALDOS HORIZOTAL sin resolver "
              "(se dejan como estaban):".format(len(dudas)))
        for d in dudas:
            print("  " + d)
    vaciar(con_formulas)
    ruta_base = os.path.join(DESTINO, "saldos_base.xlsx")
    con_formulas.save(ruta_base)
    print("escrito {} ({:.1f} MB)".format(
        ruta_base, os.path.getsize(ruta_base) / 1e6))
    return 0



# ---------------------------------------------------------------------------
# SALDOS HORIZOTAL: el mismo corrimiento, detectado por razón social
# ---------------------------------------------------------------------------

# La hoja horizontal es un pivote empresa x banco (bloque en pesos, filas 5-59, y
# bloque en dólares, 64-80) y arrastra el MISMO desfase de HSBC que SALDOS. Ahí no
# hay columna de cuenta con la que cruzar, pero sí el nombre de la empresa en la
# columna A, y varias pestañas traen la razón social del titular. Con eso alcanza.
#
# Se excluyen las pestañas que identifican por ALIAS y no por razón social ('ASKE
# MN', 'ABACOM 0875', 'PSC 0014'): ahí el nombre no se puede comparar y cualquier
# "corrección" sería adivinar.
HOJAS_CON_RAZON_SOCIAL = {"HSBC", "BANREGIO", "BANORTE", "SANTANDER", "BAJIO",
                          "INBURSA", "MULTIVA", "SABADELL", "BANAMEX"}

BLOQUES_HORIZONTAL = ((5, 59, "MXN"), (64, 80, "USD"))

_RUIDO_RAZON = re.compile(
    r"\b(S\.?A\.?P?\.?I?\.?|DE|C\.?V\.?|S\.?C\.?|A\.?C\.?|SAPI|SOFOM|ENR)\b")

_REF_SUELTA = re.compile(
    r"'?([A-Z+ ]{3,12})'?!(?:\$?([A-Z]{1,2})\$?(\d+)|#REF!)")


def _razon(texto):
    plano = _sin_acentos(str(texto or "")).upper()
    plano = re.sub(r"[^A-Z0-9 ]", " ", plano)
    plano = _RUIDO_RAZON.sub(" ", plano)
    return " ".join(plano.split())


def _mismo_titular(a, b, n=16):
    """Dos nombres son de la misma empresa.

    Se comparan los primeros caracteres porque las pestañas truncan el titular a
    lo que quepa en la columna ('TRANSPORTES Y EQUIPOS ASAMAZ S' contra
    'TRANSPORTES Y EQUIPOS ASAMAZ S.A. DE C.V.')."""
    a, b = _razon(a), _razon(b)
    if not a or not b:
        return False
    corte = min(n, len(a), len(b))
    return a[:corte] == b[:corte]


def _divisas_por_celda(renglones, inventario, desfases=()):
    """(pestaña, fila) -> divisa, según lo que declara SALDOS para ese renglón.

    Hace falta porque varias pestañas no traen columna de moneda: en HSBC el
    bloque de dólares se distingue solo porque SALDOS marca 'DLS' esos renglones.

    Se usan las filas YA CORREGIDAS. Con las originales, la última fila de un
    bloque corrido se quedaba sin divisa y quedaba fuera de las candidatas — que
    es justo la que hace falta para reparar."""
    arreglos = {(d["banda"], d["fila"]): d["ref_fila_buena"] for d in desfases}
    divisas = {}
    for r in renglones:
        marca = (r.get("marca") or "").upper()
        fila = arreglos.get((r["banda"], r["fila"]), r["ref_fila"])
        divisas[(clave_hoja(r["ref_hoja"]), fila)] = (
            "USD" if marca in ("DLS", "DLLS", "USD") else "MXN")
    for hoja, filas in inventario.items():
        for fila, datos in filas.items():
            if (hoja, fila) in divisas:
                continue
            moneda = str(datos.get("moneda") or "MXN").upper()
            divisas[(hoja, fila)] = "USD" if moneda == "USD" else "MXN"
    return divisas


def reparar_horizontal(libro, inventario, renglones, desfases=()):
    """Corrige las referencias de SALDOS HORIZOTAL que apuntan a otra empresa.

    Devuelve la lista de cambios. Solo toca una celda cuando hay UNA candidata:
    con cero se blanquea si venía rota (la empresa no tiene cuenta en ese banco y
    esa divisa) y con dos o más se deja como está y se reporta."""
    hoja = libro["SALDOS HORIZOTAL"]
    divisas = _divisas_por_celda(renglones, inventario, desfases)
    arreglos = {(d["banda"], d["fila"]): d["ref_fila_buena"] for d in desfases}
    usadas_en_saldos = {
        (clave_hoja(r["ref_hoja"]),
         arreglos.get((r["banda"], r["fila"]), r["ref_fila"]))
        for r in renglones}
    cambios, sin_resolver = [], []

    for ini, fin, divisa in BLOQUES_HORIZONTAL:
        for fila in range(ini, fin + 1):
            empresa = hoja.cell(fila, 1).value
            if not empresa or str(empresa).strip().upper() == "TOTAL":
                continue
            for col in range(3, 18):
                celda = hoja.cell(fila, col)
                formula = celda.value
                if not isinstance(formula, str) or "!" not in formula:
                    continue
                nueva, dudas = _reparar_formula(
                    formula, empresa, divisa, inventario, divisas,
                    usadas_en_saldos)
                sin_resolver.extend(
                    "{} {}".format(celda.coordinate, d) for d in dudas)
                if nueva is not None and nueva != formula:
                    cambios.append({
                        "celda": celda.coordinate, "empresa": str(empresa),
                        "divisa": divisa, "antes": formula, "despues": nueva})
                    celda.value = nueva or None
    return cambios, sin_resolver


def _reparar_formula(formula, empresa, divisa, inventario, divisas,
                     usadas_en_saldos=None):
    """Reescribe las referencias mal apuntadas de una fórmula. None si no cambia."""
    dudas = []
    ya_usadas = {(m.group(1).strip(), int(m.group(3)))
                 for m in _REF_SUELTA.finditer(formula) if m.group(3)}

    def reemplazo(m):
        nombre = m.group(1).strip()
        if nombre not in HOJAS_CON_RAZON_SOCIAL:
            return m.group(0)
        roto = m.group(3) is None
        fila_actual = None if roto else int(m.group(3))
        if not roto:
            datos = inventario.get(nombre, {}).get(fila_actual, {})
            titular = datos.get("titular") or datos.get("alias")
            if _mismo_titular(titular, empresa):
                return m.group(0)
            if not titular:
                # Fila vacía: puede ser que ese día el portal no la trajera.
                # No es prueba de desfase.
                return m.group(0)
        candidatas = [
            f for f, d in inventario.get(nombre, {}).items()
            if _mismo_titular(d.get("titular") or d.get("alias"), empresa)
            and divisas.get((nombre, f)) == divisa
            and (f == fila_actual or (nombre, f) not in ya_usadas)
        ]
        if len(candidatas) > 1 and usadas_en_saldos is not None:
            # La hoja horizontal debe reflejar las MISMAS cuentas que SALDOS. Si
            # varias filas del banco son de esa empresa, gana la que SALDOS ya
            # lleva: las otras son cuentas que el formato no contempla.
            preferidas = [f for f in candidatas
                          if (nombre, f) in usadas_en_saldos]
            if len(preferidas) == 1:
                candidatas = preferidas
        if len(candidatas) == 1:
            col = m.group(2) or _columna_tipica(nombre, inventario)
            return "{}!{}{}".format(_nombre_ref(nombre), col, candidatas[0])
        if roto and not candidatas:
            # La empresa no tiene cuenta en ese banco y esa divisa: la referencia
            # rota no ocultaba nada, sobraba.
            return "0"
        dudas.append("{}: {} candidata(s)".format(nombre, len(candidatas)))
        return m.group(0)

    nueva = _REF_SUELTA.sub(reemplazo, formula)
    return (nueva if nueva != formula else None), dudas


def _nombre_ref(nombre):
    return "'{}'".format(nombre) if " " in nombre or "+" in nombre else nombre


def _columna_tipica(nombre, inventario):
    """Columna de saldo de esa pestaña, para rearmar una referencia rota."""
    for regiones in (REGIONES.get(nombre) or ()):
        if "saldo" in regiones.cols:
            return regiones.cols["saldo"]
    return "A"

if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

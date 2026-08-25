"""Fase 0 del módulo Saldos: qué cuentas del formato NO están en el catálogo.

El reporte diario de saldos se arma hoy en `FORMATO DE SALDOS TESORERIA.xlsx`,
donde cada saldo es una fórmula que apunta a una celda FIJA de la hoja pegada del
banco (`=BANAMEX!E4`). Esa relación cuenta↔empresa no está escrita en ningún lado:
vive en la posición de la fila, y por eso se rompe en silencio cuando el portal
cambia el orden de sus filas.

Para sustituirla por una búsqueda por número de cuenta hace falta que TODAS esas
cuentas estén en `CUENTAS DISPERSION.xlsx`. Este script cruza ambas fuentes y emite
un Excel con las que faltan, listo para que tesorería lo complete con la CLABE y el
número de cuenta.

Es de UN SOLO USO (no lo consume la app). Se corre así:

    python scripts/saldos_cuentas_faltantes.py [ruta del formato] [-o salida.xlsx]
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from core import cuentas_dispersion
from core.catalogo_bancos import CATALOGO_BANCOS

# Ruta por defecto del formato que usa tesorería.
_FORMATO = os.path.join("descargas", "ReporteCuentas",
                        "FORMATO DE SALDOS TESORERIA.xlsx")
_HOJA = "SALDOS"
# La hoja llega hasta la fila 237; se recorre completa.
_ULTIMA_FILA = 237

# Una fórmula de saldo: '=BANAMEX!E4', "=+'BX+ SCO'!G3", '=$I$35'… Solo interesan
# las que apuntan a OTRA hoja (las de la propia hoja son sumas y totales).
_RE_REF = re.compile(r"^=\+?'?([A-ZÉÁ+ ]+)'?!\$?([A-Z]{1,2})\$?(\d+)$")

# Nombre de banco del formato -> prefijo de CLABE. El formato los escribe a mano,
# con typos ('Scotianbank', 'Santandeer'), sufijos ('BBVA TDE', 'HSBC REF',
# 'Monex (Cred)') y abreviaturas ('Bx+', 'Banbajio'), así que se compara por
# PREFIJO del nombre ya normalizado a mayúsculas sin acentos ni signos.
_PREFIJO_BANCO = {
    "BANAMEX": "002",
    "BANBAJIO": "030",
    "BAJIO": "030",
    "BANCOPPEL": "137",
    "BANORTE": "072",
    "BANREGIO": "058",
    "BBVA": "012",
    "BANCOMER": "012",
    "BX": "113",          # 'Bx+' -> Ve por Más
    "HSBC": "021",
    "INBURSA": "036",
    "INTERCAM": "630",
    "MONEX": "112",
    "MULTIVA": "132",
    "SABADELL": "156",
    "SANTANDER": "014",
    "SANTANDEER": "014",  # typo del formato
    "SCOTIABANK": "044",
    "SCOTIANBANK": "044",  # typo del formato
}


def _sin_acentos(texto) -> str:
    """Mayúsculas, sin acentos y sin nada que no sea letra o dígito."""
    plano = unicodedata.normalize("NFKD", str(texto or ""))
    plano = plano.encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]", "", plano.upper())


def prefijo_de_banco(nombre: str) -> str | None:
    """Prefijo de CLABE del banco que nombra `nombre`, o None si no se reconoce.

    Se busca por prefijo y de más largo a más corto para que 'BANBAJIO' gane sobre
    'BAJIO' y 'BANCOMER' no se confunda con 'BANCOPPEL'."""
    plano = _sin_acentos(nombre)
    if not plano:
        return None
    for clave in sorted(_PREFIJO_BANCO, key=len, reverse=True):
        if plano.startswith(clave):
            return _PREFIJO_BANCO[clave]
    return None


def _celdas(ruta: str, hoja: str) -> dict:
    """La hoja como {fila: [valores por columna]}, con las FÓRMULAS sin evaluar."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=False)
    try:
        ws = wb[hoja]
        out = {}
        for r, fila in enumerate(
                ws.iter_rows(min_row=1, max_row=_ULTIMA_FILA, values_only=True), 1):
            out[r] = list(fila) + [None] * (24 - len(fila))
        return out
    finally:
        wb.close()


def _region(celdas: dict, col_empresa: str, col_banco: str,
            col_cuenta: str, col_saldo: str) -> list[dict]:
    """Cuentas de una de las dos regiones de la hoja SALDOS.

    Una fila es CUENTA si su columna de saldo trae una referencia a la hoja de un
    banco. Es ENCABEZADO DE EMPRESA si la columna de empresa trae texto y no hay
    saldo. Se recorre de arriba abajo arrastrando la última empresa vista, que es
    exactamente como está organizada la hoja."""
    def col(fila: int, letra: str):
        return celdas[fila][ord(letra) - 65]

    cuentas: list[dict] = []
    empresa = "(sin empresa)"
    for r in range(1, _ULTIMA_FILA + 1):
        saldo = col(r, col_saldo)
        texto_saldo = str(saldo) if saldo is not None else ""
        ref = _RE_REF.match(texto_saldo.strip())
        if ref:
            cuentas.append({
                "empresa": empresa,
                "banco": str(col(r, col_banco) or "").strip(),
                "cuenta4": re.sub(r"\D", "", str(col(r, col_cuenta) or "")),
                "origen": f"{ref.group(1)}!{ref.group(2)}{ref.group(3)}",
                "fila": r,
            })
            continue
        emp = col(r, col_empresa)
        # Encabezado de empresa: texto, no fórmula y con cuerpo suficiente (evita
        # tomar por empresa las etiquetas sueltas de una letra o dos).
        if (isinstance(emp, str) and emp.strip() and len(emp.strip()) > 3
                and not emp.startswith("=") and not texto_saldo.startswith("=")):
            empresa = emp.strip()
    return cuentas


def leer_formato(ruta: str) -> list[dict]:
    """Las cuentas mapeadas en la hoja SALDOS, de sus dos regiones.

    La izquierda (A/B/I) y la derecha (Q/R/U) tienen la misma forma pero distinta
    posición; en los bloques 'BANREGIO CTA NUEVAS' y 'BBVA CUENTAS NUEVAS' de la
    derecha están invertidas (agrupa por banco y la columna Q trae la empresa), así
    que ahí 'empresa' y 'banco' salen cambiados. Se marcan al exportar."""
    celdas = _celdas(ruta, _HOJA)
    izquierda = _region(celdas, "F", "A", "B", "I")
    derecha = _region(celdas, "Q", "Q", "R", "U")
    for c in izquierda:
        c["region"] = "izquierda (A/B/I)"
    for c in derecha:
        c["region"] = "derecha (Q/R/U)"
    return izquierda + derecha


def indice_catalogo(catalogo) -> dict:
    """Índice del catálogo de dispersión por (prefijo de banco, últimos 4 dígitos).

    De cada cuenta se derivan VARIAS colas porque el mismo número aparece escrito de
    formas distintas: el numeroCuenta pelón, la CLABE completa y la cuenta que la
    CLABE lleva embebida (posiciones 7 a 17). Ojo: los últimos 4 de una CLABE NO son
    los de la cuenta, porque el dígito verificador va al final."""
    idx: dict = {}
    for id_empresa in catalogo.empresas():
        for reg in catalogo._registros(id_empresa):
            clabe = re.sub(r"\D", "", reg.get("clabe", ""))
            numero = re.sub(r"\D", "", reg.get("numero", ""))
            prefijo = clabe[:3] if len(clabe) == 18 else None
            colas = {numero[-4:]} if len(numero) >= 4 else set()
            if len(clabe) == 18:
                colas.add(clabe[-4:])
                colas.add(clabe[6:17][-4:])
            for cola in colas:
                if cola:
                    idx.setdefault((prefijo, cola), []).append((id_empresa, reg))
    return idx


def enderezar_invertidas(cuentas: list[dict]) -> int:
    """Corrige los bloques que vienen al revés y devuelve cuántos arregló.

    'BANREGIO CTA NUEVAS' y 'BBVA CUENTAS NUEVAS' agrupan por BANCO en vez de por
    empresa: el encabezado del bloque nombra al banco y la columna que en el resto
    de la hoja lleva el banco trae aquí a la empresa. Se detecta por el dato, no por
    el título: si el encabezado del bloque se reconoce como banco y el contenido de
    la columna de banco NO, están cambiados."""
    arreglados = 0
    for c in cuentas:
        if prefijo_de_banco(c["empresa"]) and not prefijo_de_banco(c["banco"]):
            c["empresa"], c["banco"] = c["banco"], c["empresa"]
            c["invertida"] = True
            arreglados += 1
    return arreglados


def cruzar(cuentas: list[dict], idx: dict) -> list[dict]:
    """Marca cada cuenta del formato como encontrada o faltante, con el motivo."""
    enderezar_invertidas(cuentas)
    for c in cuentas:
        prefijo = prefijo_de_banco(c["banco"])
        cola = c["cuenta4"][-4:].zfill(4) if c["cuenta4"] else ""
        c["prefijo"] = prefijo or ""
        c["banco_catalogo"] = CATALOGO_BANCOS.get(prefijo or "", "")
        if not cola:
            c["estado"], c["motivo"] = "FALTA", "la fila no trae dígitos de cuenta"
        elif prefijo is None:
            c["estado"] = "FALTA"
            c["motivo"] = "no se reconoce el banco (¿la columna trae la empresa?)"
        elif (prefijo, cola) in idx:
            c["estado"] = "OK"
            ide, reg = idx[(prefijo, cola)][0]
            c["motivo"] = f"id {ide} · {reg.get('cuenta', '')}"
            # Se copian los datos del catálogo para que la hoja de referencia sirva
            # de ejemplo de cómo llenar las que faltan.
            c["clabe"] = reg.get("clabe", "")
            c["numero"] = reg.get("numero", "")
            c["id_empresa"] = ide
        else:
            c["estado"] = "FALTA"
            c["motivo"] = "no está en CUENTAS DISPERSION"
    return cuentas


_ENCABEZADOS = [
    ("Empresa (formato)", 38), ("Banco (formato)", 20), ("Últimos 4", 12),
    ("Banco por CLABE", 22), ("Prefijo", 10), ("CLABE", 22),
    ("numeroCuenta", 18), ("id Empresa", 12),
    ("Motivo", 40), ("Región", 20), ("Celda origen", 16), ("Fila", 8),
]
_AZUL = "FF317FB1"


def _hoja(wb, titulo: str, cuentas: list[dict], nota: str):
    ws = wb.create_sheet(titulo)
    ws["A1"] = nota
    ws["A1"].font = Font(name="Arial", size=11, bold=True)
    for i, (etq, ancho) in enumerate(_ENCABEZADOS, 1):
        celda = ws.cell(3, i, etq)
        celda.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[celda.column_letter].width = ancho
    for f, c in enumerate(cuentas, 4):
        # En las FALTANTES estas tres van vacías a propósito: son justo las que
        # tesorería tiene que llenar. En las que ya están, se rellenan con lo del
        # catálogo para que sirvan de ejemplo del formato esperado.
        valores = [c["empresa"], c["banco"], c["cuenta4"], c["banco_catalogo"],
                   c["prefijo"], c.get("clabe", ""), c.get("numero", ""),
                   c.get("id_empresa", ""), c["motivo"], c["region"],
                   c["origen"], c["fila"]]
        for i, v in enumerate(valores, 1):
            celda = ws.cell(f, i, v)
            celda.font = Font(name="Arial", size=10)
            # Los números de cuenta van como TEXTO o Excel se come los ceros a la
            # izquierda ('0454' no es 454, y una CLABE que empieza en 0 tampoco).
            if i in (3, 6, 7):
                celda.number_format = "@"
    ws.freeze_panes = "A4"
    return ws


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("formato", nargs="?", default=_FORMATO,
                   help="ruta de FORMATO DE SALDOS TESORERIA.xlsx")
    p.add_argument("-o", "--salida", default="CUENTAS SALDOS - por completar.xlsx")
    args = p.parse_args()

    if not os.path.exists(args.formato):
        print(f"No se encontró el formato: {args.formato}")
        return 1

    cuentas = leer_formato(args.formato)
    catalogo = cuentas_dispersion.CatalogoCuentasDispersion()
    if not catalogo.disponible():
        print("El catálogo CUENTAS DISPERSION.xlsx no está instalado.")
        return 1
    cruzar(cuentas, indice_catalogo(catalogo))

    faltan = [c for c in cuentas if c["estado"] == "FALTA"]
    ok = [c for c in cuentas if c["estado"] == "OK"]
    total_cat = sum(len(v) for v in catalogo.datos.values())

    print(f"Formato   : {args.formato}")
    print(f"Catálogo  : {total_cat} cuentas en {len(catalogo.empresas())} empresas")
    print()
    print(f"  cuentas mapeadas en el formato : {len(cuentas):>4}")
    print(f"  ya están en el catálogo        : {len(ok):>4} "
          f"({len(ok) * 100 // max(1, len(cuentas))} %)")
    print(f"  FALTAN                         : {len(faltan):>4}")
    print()
    for motivo in sorted({c["motivo"] for c in faltan}):
        print(f"    {sum(1 for c in faltan if c['motivo'] == motivo):>3}  {motivo}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja(wb, "Por completar", faltan,
          "Cuentas del reporte de saldos que NO están en CUENTAS DISPERSION. "
          "Llena CLABE, numeroCuenta e id Empresa y pásalas al catálogo.")
    _hoja(wb, "Ya en el catálogo", ok,
          "Cuentas del reporte de saldos que ya están en el catálogo "
          "(referencia, no requieren acción).")
    wb.save(args.salida)
    print(f"\nGenerado: {args.salida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

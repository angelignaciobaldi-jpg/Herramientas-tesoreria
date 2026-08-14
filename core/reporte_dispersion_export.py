"""Generación del reporte XLSX de solicitudes SELECCIONADAS para dispersar.

Es el espejo de `core.reporte_dispersion` (que LEE los reportes del SIPP): aquí se
ESCRIBE un Excel con el mismo layout que el que descarga el SIPP ("Generar XLS"),
para que el usuario pueda revisar en un archivo familiar lo que eligió pagar.

  - Una HOJA por empresa (las solicitudes seleccionadas de cada empresa).
  - Título, bloque de filtros (B3:G6) y encabezados como el reporte original; la
    única diferencia en los filtros es que el campo "Sucursal" se reemplaza por
    "Fecha Vencimiento".
  - Filas de datos agrupadas por cuenta bancaria y, al cierre de cada grupo, una
    fila 'TOTAL PROGRAMADO' con la suma del Saldo Programado del grupo. Las notas de
    crédito aparecen en su renglón en negativo y RESTAN de ese total, de modo que
    coincide con el importe del TXT y con el neto que muestra el SIPP.

No depende de Flet: opera sobre FilaSolicitud, por lo que es fácilmente testeable.
"""

from __future__ import annotations

import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from core.reporte_dispersion import FilaSolicitud, total_a_pagar

# Colores del reporte original del SIPP (título/encabezado azul, total gris).
_AZUL = "FF317FB1"
_GRIS_TOTAL = "FFE0E0E0"
_FUENTE = "Arial"

# Encabezados de la tabla, en orden (las 15 columnas del reporte del SIPP más
# 'Equiv. MXN', que es propia de la app: el equivalente en pesos de los proveedores
# USD marcados 'pagar en pesos').
_ENCABEZADOS = [
    "Folio", "Tipo de factura", "Folio Factura", "Empresa", "Proveedor",
    "Fecha Factura", "Fecha Vencimiento", "Tipo Solicitud", "Moneda", "Producto",
    "Total Factura", "Saldo Factura", "Saldo Programado", "Cuenta Bancaria",
    "Comentarios", "Equiv. MXN",
]
# Columnas (1-based) con formato de moneda: Total/Saldo Factura, Saldo Programado y
# el equivalente en MXN.
_COLS_MONTO = {11, 12, 13, 16}
# Fila 'TOTAL PROGRAMADO' del cierre de grupo: etiqueta en 'Cuenta Bancaria' (N) y
# el importe en 'Comentarios' (O), tal como lo coloca el reporte del SIPP.
_COL_TOTAL_ETIQUETA = 14
_COL_TOTAL_VALOR = 15
_COL_EQUIV_MXN = 16
_FMT_MONEDA = '"$"#,##0.00'

# Anchos de columna (A..P), copiados del reporte original para que se vea igual.
_ANCHOS = {
    "A": 13.33, "B": 8.33, "C": 16.66, "D": 16.66, "E": 46.66, "F": 13.33,
    "G": 13.33, "H": 13.33, "I": 13.33, "J": 20.0, "K": 13.33, "L": 13.33,
    "M": 13.33, "N": 41.66, "O": 41.66, "P": 15.0,
}

# Caracteres no permitidos por Excel en el nombre de una hoja.
_INVALIDOS_HOJA = re.compile(r"[\[\]:*?/\\]")


def _monto(valor) -> float:
    try:
        return float(valor or 0)
    except (TypeError, ValueError):
        return 0.0


def _nombre_hoja(empresa: str, usados: set[str]) -> str:
    """Nombre de hoja válido y único: sin caracteres prohibidos y <= 31 chars."""
    base = _INVALIDOS_HOJA.sub(" ", (empresa or "Sin empresa").strip())[:31] or "Hoja"
    nombre = base
    i = 2
    while nombre.lower() in usados:
        sufijo = f" ({i})"
        nombre = base[: 31 - len(sufijo)] + sufijo
        i += 1
    usados.add(nombre.lower())
    return nombre


def _etiqueta(ws, coord: str, texto: str) -> None:
    c = ws[coord]
    c.value = texto
    c.font = Font(name=_FUENTE, size=10, bold=True)
    c.alignment = Alignment(horizontal="right")


def _valor(ws, coord: str, texto: str) -> None:
    c = ws[coord]
    c.value = texto
    c.font = Font(name=_FUENTE, size=10)
    c.alignment = Alignment(horizontal="left")


def _un_valor(filas: list[FilaSolicitud], campo: str, plural: str) -> str:
    """Si TODAS las filas comparten un mismo valor del campo, lo devuelve; si hay
    más de uno (o ninguno), devuelve `plural` (p. ej. 'Todos' / 'Todas')."""
    distintos = {(getattr(f, campo, "") or "").strip() for f in filas}
    distintos.discard("")
    return next(iter(distintos)) if len(distintos) == 1 else plural


def _bloque_filtros(ws, empresa: str, filas: list[FilaSolicitud], filtros: dict,
                    con_tipo_cambio: bool = False) -> None:
    """Rellena el área de metadatos B3:G6 (etiqueta a la izquierda, valor a la
    derecha), igual que el reporte del SIPP pero con 'Fecha Vencimiento' en lugar
    de 'Sucursal'. Con `con_tipo_cambio` añade el TC en la fila 7."""
    ws.merge_cells("C4:D4")
    ws.merge_cells("G4:H4")
    # Columna izquierda.
    _etiqueta(ws, "B3", "Empresa:");          _valor(ws, "C3", empresa)
    _etiqueta(ws, "B4", "Proveedor:")
    _valor(ws, "C4", _un_valor(filas, "proveedor", "Todos"))
    _etiqueta(ws, "B5", "Cuenta Bancaria:")
    _valor(ws, "C5", _un_valor(filas, "cuenta_bancaria", "Todas"))
    _etiqueta(ws, "B6", "Fecha Inicio:");     _valor(ws, "C6", filtros.get("fecha_inicio", ""))
    # Columna derecha ('Sucursal' -> 'Fecha Vencimiento').
    _etiqueta(ws, "F3", "Fecha Vencimiento:")
    _valor(ws, "G3", filtros.get("fecha_vencimiento", "N/A"))
    _etiqueta(ws, "F4", "Folio:");            _valor(ws, "G4", filtros.get("folio", "Todos"))
    _etiqueta(ws, "F5", "Tipo Solicitud:")
    _valor(ws, "G5", filtros.get("tipo_solicitud", "Todos"))
    _etiqueta(ws, "F6", "Fecha Fin:");        _valor(ws, "G6", filtros.get("fecha_fin", ""))
    # Fila 7: tipo de cambio. Solo en las hojas con proveedores marcados 'pagar en
    # pesos' (es con el que se calcula 'Equiv. MXN'); en las demás la fila queda en
    # blanco y el reporte se ve como el del SIPP.
    if con_tipo_cambio and filtros.get("tipo_cambio"):
        _etiqueta(ws, "B7", "Tipo de cambio:")
        _valor(ws, "C7", filtros["tipo_cambio"])


def _fila_datos(f: FilaSolicitud, equiv_mxn: float | None = None,
                con_equiv: bool = False) -> list:
    """Valores de una fila en el orden de _ENCABEZADOS.

    `con_equiv` decide si la fila incluye la columna 'Equiv. MXN'; se omite en las
    hojas sin proveedores en pesos, para que salgan idénticas al reporte del SIPP.
    `equiv_mxn` es el Saldo Programado convertido a pesos: solo lo traen las filas de
    proveedores marcados 'pagar en pesos'. En el resto de una hoja que SÍ lleva la
    columna, la celda queda vacía (None), que es más honesto que un 0: significa 'no
    aplica', no 'cero pesos'."""
    valores = [
        f.folio, f.tipo, f.folio_factura, f.empresa, f.proveedor,
        f.fecha_factura, f.fecha_vencimiento, f.tipo_solicitud, f.moneda, f.producto,
        _monto(f.total_factura), _monto(f.saldo_factura), _monto(f.saldo_programado),
        f.cuenta_bancaria, f.comentarios,
    ]
    if con_equiv:
        valores.append(None if equiv_mxn is None else round(equiv_mxn, 2))
    return valores


def _construir_hoja(ws, empresa: str, filas: list[FilaSolicitud], filtros: dict,
                    pares_pesos: set | None = None, tc: float | None = None) -> None:
    # --- Título (A1:G1) ---
    ws.merge_cells("A1:G1")
    tit = ws["A1"]
    tit.value = "Solicitudes de Pago a Proveedores (NO PEMEX)"
    tit.font = Font(name=_FUENTE, size=22, bold=True, color=_AZUL)
    tit.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22.5
    # NOTA: la celda J1 (impreso por / fecha) se deja VACÍA a propósito.

    # El tipo de cambio y la columna 'Equiv. MXN' solo aparecen en las hojas con al
    # menos un proveedor marcado 'pagar en pesos' (y con tipo de cambio disponible).
    # En las demás no aportan nada y el reporte sale idéntico al del SIPP.
    pares_pesos = pares_pesos or set()
    con_equiv = bool(pares_pesos) and bool(tc)
    encabezados = _ENCABEZADOS if con_equiv else _ENCABEZADOS[:-1]

    # --- Bloque de filtros (B3:G6, más el TC en la 7 si aplica) ---
    _bloque_filtros(ws, empresa, filas, filtros, con_equiv)

    # --- Encabezados de la tabla (fila 9) ---
    # Van una fila más abajo que en el reporte del SIPP: el bloque de filtros ocupa
    # hasta la 7 (el tipo de cambio) y se deja la 8 en blanco como separación.
    fila_enc = 9
    for col, titulo in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_enc, column=col, value=titulo)
        c.font = Font(name=_FUENTE, size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=_AZUL)
        c.alignment = Alignment(horizontal="center")

    # --- Datos agrupados por cuenta bancaria, con TOTAL PROGRAMADO por grupo ---
    grupos: dict[str, list[FilaSolicitud]] = {}
    orden: list[str] = []
    for f in filas:
        if f.cuenta_bancaria not in grupos:
            grupos[f.cuenta_bancaria] = []
            orden.append(f.cuenta_bancaria)
        grupos[f.cuenta_bancaria].append(f)

    fila = fila_enc + 1
    for cuenta in orden:
        grupo = grupos[cuenta]
        for f in grupo:
            # Equivalente en MXN solo para los pares marcados 'pagar en pesos'; el
            # resto de la hoja va vacío.
            equiv = None
            if con_equiv and (f.proveedor, f.cuenta_bancaria) in pares_pesos:
                equiv = (f.saldo_programado or 0) * tc
            for col, valor in enumerate(
                    _fila_datos(f, equiv, con_equiv), start=1):
                c = ws.cell(row=fila, column=col, value=valor)
                c.font = Font(name=_FUENTE, size=10)
                if col in _COLS_MONTO:
                    c.alignment = Alignment(horizontal="right")
                    c.number_format = _FMT_MONEDA
                else:
                    c.alignment = Alignment(horizontal="center")
            fila += 1
        # Fila TOTAL PROGRAMADO del grupo.
        etq = ws.cell(row=fila, column=_COL_TOTAL_ETIQUETA, value="TOTAL PROGRAMADO")
        etq.font = Font(name=_FUENTE, size=10, bold=True)
        etq.alignment = Alignment(horizontal="right")
        # Con las notas de crédito descontadas: el archivo refleja lo que se va a
        # PAGAR, así que su total debe ser el mismo del TXT (ver total_a_pagar).
        tot = ws.cell(row=fila, column=_COL_TOTAL_VALOR,
                      value=total_a_pagar(grupo))
        tot.font = Font(name=_FUENTE, size=10, bold=True)
        tot.alignment = Alignment(horizontal="right")
        tot.number_format = _FMT_MONEDA
        tot.fill = PatternFill("solid", fgColor=_GRIS_TOTAL)
        fila += 2  # una fila en blanco de separación antes del siguiente grupo

    # --- Ancho de columnas y paneles fijos (título + filtros + encabezado) ---
    for col, ancho in _ANCHOS.items():
        if col == "P" and not con_equiv:
            continue  # esa columna no existe en esta hoja
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A10"


def generar(ruta: str, seleccion: dict[str, list[FilaSolicitud]], filtros: dict,
            pares_pesos: dict | None = None, tc: float | None = None) -> None:
    """Crea el reporte XLSX en `ruta`, con una hoja por grupo empresa+moneda.

    Args:
        ruta: ruta destino .xlsx.
        seleccion: {clave: [FilaSolicitud seleccionadas]} donde `clave` es el grupo
            'Empresa - Moneda' (una hoja por clave). El nombre de la hoja usa la
            clave; la celda 'Empresa:' del bloque de filtros usa la empresa real
            de las filas.
        filtros: valores del bloque de metadatos ya resueltos por la UI:
            fecha_inicio, fecha_fin, fecha_vencimiento ('N/A' si no aplica),
            folio ('Todos' si no aplica), tipo_solicitud ('Todos' o concatenado) y,
            opcional, tipo_cambio (texto ya formateado; si falta, no se imprime).
        pares_pesos: {clave: {(proveedor, cuenta), ...}} de los proveedores marcados
            'pagar en pesos'. Solo esas filas llevan 'Equiv. MXN'.
        tc: tipo de cambio con el que se calcula esa columna. Sin él (o sin pares),
            la columna queda vacía y el reporte se ve como el del SIPP.
    """
    pares_pesos = pares_pesos or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # se crea una hoja por grupo; se quita la de por defecto
    usados: set[str] = set()
    for clave, filas in seleccion.items():
        ws = wb.create_sheet(title=_nombre_hoja(clave, usados))
        empresa = filas[0].empresa if filas else clave
        _construir_hoja(ws, empresa, filas, filtros,
                        pares_pesos.get(clave) or set(), tc)
    if not wb.sheetnames:  # sin selección: deja un libro válido (no debería pasar)
        wb.create_sheet(title="Sin datos")
    wb.save(ruta)

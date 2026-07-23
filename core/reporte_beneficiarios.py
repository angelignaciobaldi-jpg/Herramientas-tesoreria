"""Reporte Excel del apartado de Alta de beneficiarios.

Exporta las columnas del listado SIN 'Estado' ni 'Acciones' (no aplican a un
reporte): CLABE, Monto, Banco, Tipo, Beneficiario, Alias, Email.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_AZUL = "1F4E78"
_GRIS = "D9D9D9"
_BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def generar(ruta: str, registros: list[dict]) -> None:
    """Crea el Excel del reporte.

    Args:
        ruta: ruta destino .xlsx
        registros: lista de dicts con claves: clabe, monto (float|None), banco,
            tipo, beneficiario, alias, email.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beneficiarios"

    ws["A1"] = "Reporte de beneficiarios"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)

    fila = 3
    encabezados = ["#", "CLABE", "Monto", "Banco", "Tipo",
                   "Beneficiario", "Alias", "Email"]
    for col, titulo in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDE

    CENTRADAS = {1, 2, 4, 5}   # #, CLABE, Banco, Tipo
    COL_MONTO = 3
    ini = fila + 1
    for i, r in enumerate(registros, start=1):
        monto = r.get("monto")
        valores = [
            i,
            r.get("clabe", ""),
            float(monto) if monto is not None else None,
            r.get("banco", ""),
            r.get("tipo", ""),
            r.get("beneficiario", ""),
            r.get("alias", ""),
            r.get("email", ""),
        ]
        for col, valor in enumerate(valores, start=1):
            c = ws.cell(row=ini + i - 1, column=col, value=valor)
            c.border = _BORDE
            if col in CENTRADAS:
                c.alignment = Alignment(horizontal="center")
            if col == COL_MONTO and valor is not None:
                c.number_format = '#,##0.00'

    # Total de montos (si hay al menos un registro).
    if registros:
        fila_total = ini + len(registros)
        ws.cell(row=fila_total, column=COL_MONTO - 1, value="TOTAL").font = Font(bold=True)
        ct = ws.cell(row=fila_total, column=COL_MONTO,
                     value=f"=SUM(C{ini}:C{fila_total - 1})")
        ct.font = Font(bold=True)
        ct.number_format = '#,##0.00'
        ct.fill = PatternFill("solid", fgColor=_GRIS)

    anchos = {"A": 5, "B": 22, "C": 14, "D": 16, "E": 24, "F": 30, "G": 30, "H": 28}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(ruta)

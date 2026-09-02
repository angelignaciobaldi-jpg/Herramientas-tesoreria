"""Catálogo de 'Cuentas de dispersión' (por id de empresa).

Lee un Excel SENCILLO para filtrar las cuentas en la pantalla de Dispersión
(No Pemex):

  - ``id Empresa``          -> FK al id de la empresa (ver EMPRESAS en la pantalla).
  - ``Cuenta``              -> la cuenta que se MUESTRA en el selector y por la que
                               el RPA busca la cuenta.
  - ``numeroCuenta``        -> número de cuenta "pelón" (opcional). Sirve para casar
                               un comprobante cuando el nombre de la cuenta no trae
                               dígitos: hay cuentas con nombres como 'PETRO SMART
                               HERMOSILLO BBVA', y el comprobante del banco las
                               identifica por su número.
  - ``CLABE interbancaria`` -> la CLABE de origen que se usa como cuenta origen del
                               TXT en pesos (opcional; se muestra en su propio
                               selector).
  - ``moneda``              -> id del catálogo MONEDAS (1 = Pesos, 2 = Dólar…).
                               Acota qué cuentas se ofrecen: las solicitudes ya
                               vienen separadas por moneda, y no se paga una en
                               dólares desde una cuenta en pesos. Opcional; una
                               cuenta sin moneda se sigue ofreciendo para todas
                               (ver `_de_la_moneda`).

Solo ``id Empresa`` y ``Cuenta`` son obligatorias: los Excel que no traigan las otras
siguen siendo válidos y esas columnas quedan vacías.

Es el complemento por-empresa del catálogo bancario general (`cuentas_bancarias`).
Se consulta por id de empresa (`cuentas_por_id_empresa`, `clabes_por_id_empresa`).

Las columnas se localizan por ENCABEZADO (fila 1), no por posición. Si el Excel se
actualiza, basta reabrir la app (o readjuntarlo en Configuración) para reflejarlo.
"""

from __future__ import annotations

import json
import os
import re

try:
    import openpyxl
except ImportError:  # openpyxl es opcional; sin él, el catálogo queda vacío.
    openpyxl = None

from . import instalador_catalogo, rutas
from .exportador_devoluciones import banco_formato
from .extractores import validar_clabe

# El Excel lo actualiza el usuario, así que va junto al .exe (no empaquetado).
RUTA_EXCEL = os.path.join(rutas.DATOS, "Cuentas dispersion", "CUENTAS DISPERSION.xlsx")
# Copia en caché (permite seguir trabajando aunque el Excel esté abierto/bloqueado).
# El nombre lleva versión: al cambiar la FORMA del registro (p. ej. al sumar
# 'numero') un caché viejo devolvería registros incompletos, así que se cambia el
# archivo y el anterior simplemente se ignora.
_RUTA_CACHE = os.path.join(rutas.DATOS, "_cuentas_dispersion_cache_v3.json")

# Encabezados aceptados (normalizados) para cada columna.
_HDR_ID = ("id empresa",)
_HDR_CUENTA = ("cuenta",)
# Número de cuenta (opcional): respaldo para casar comprobantes cuando el nombre de
# la cuenta no trae dígitos.
_HDR_NUMERO = ("numerocuenta", "numero cuenta", "numero de cuenta",
               "num cuenta", "no cuenta", "no. cuenta")
# CLABE interbancaria (opcional): cuenta origen del TXT en pesos.
_HDR_CLABE = ("clabe interbancaria", "clabe")
# Moneda de la cuenta (opcional): id del catálogo MONEDAS. Filtra qué cuentas se
# ofrecen para cada solicitud, que ya viene separada por moneda.
_HDR_MONEDA = ("moneda", "id moneda", "idmoneda", "tipo moneda",
               "tipo de moneda")

# Catálogo de monedas del SIPP: {id: nombre}.
MONEDAS = {
    1: "Pesos",
    2: "Dólar",
    3: "Euro",
    4: "Dólar Canadiense",
    5: "Libra esterlina",
    6: "Yen japonés",
}
MONEDA_PESOS = 1

# Siglas con que puede llegar la moneda de una solicitud -> id del catálogo. La
# solicitud trae la moneda ya normalizada (ver reporte_dispersion.normalizar_moneda:
# mayúsculas, sin puntos, MN->MXN), pero el SIPP no siempre usa el código ISO, así
# que se aceptan también las formas que se han visto en los reportes.
_ID_POR_SIGLAS = {
    "MXN": 1, "MN": 1, "PESOS": 1, "PESO": 1,
    "USD": 2, "DLLS": 2, "DLL": 2, "DOLAR": 2, "DÓLAR": 2, "DOLARES": 2,
    "EUR": 3, "EURO": 3, "EUROS": 3,
    "CAD": 4,
    "GBP": 5,
    "JPY": 6,
}


def id_moneda(moneda) -> int | None:
    """Id del catálogo MONEDAS a partir de un id (int) o de las siglas de la moneda
    ('USD', 'MXN'…). None si no se reconoce o viene vacío: quien filtra lo toma como
    'sin filtro' y muestra todas las cuentas, que es lo seguro cuando no se sabe."""
    if moneda is None or moneda == "":
        return None
    if isinstance(moneda, bool):
        return None
    if isinstance(moneda, (int, float)):
        ident = int(moneda)
        return ident if ident in MONEDAS else None
    crudo = str(moneda).strip()
    # El id puede venir como TEXTO ('2'): las columnas del Excel suelen formatearse
    # como texto para que no se pierdan los ceros a la izquierda de las cuentas, y
    # entonces el id se guarda igual. Sin esto la fila quedaría 'sin moneda' en
    # silencio y la cuenta aparecería en todas las pestañas.
    if crudo.isdigit():
        ident = int(crudo)
        return ident if ident in MONEDAS else None
    siglas = re.sub(r"[.\s]+", "", crudo).upper()
    return _ID_POR_SIGLAS.get(siglas)


class ExcelCuentasDispersionInvalido(ValueError):
    """El Excel adjuntado no tiene el formato esperado (id Empresa + Cuenta)."""


def hay_excel() -> bool:
    """True si ya hay un Excel de cuentas de dispersión colocado para consulta."""
    return os.path.exists(RUTA_EXCEL)


def eliminar_excel() -> bool:
    """Elimina el Excel instalado y su caché. Devuelve True si había algo que borrar.

    Borrar el caché junto con el Excel no es opcional: si quedara, la próxima
    consulta lo leería y el catálogo seguiría "disponible" con los datos del
    archivo que se acaba de eliminar.

    Propaga `OSError` si el archivo no se puede borrar —lo típico es tenerlo
    abierto en Excel— para que la pantalla lo explique en vez de decir que se
    borró algo que sigue ahí.
    """
    habia = os.path.exists(RUTA_EXCEL)
    if habia:
        os.remove(RUTA_EXCEL)
    try:
        if os.path.exists(_RUTA_CACHE):
            os.remove(_RUTA_CACHE)
    except OSError:
        pass  # el caché sin su Excel se regenera solo; no vale fallar por esto
    return habia


def _norm(valor) -> str:
    """Normaliza un encabezado: minúsculas, sin espacios/underscores repetidos."""
    return re.sub(r"[\s_]+", " ", str(valor or "").strip().casefold())


def _id_empresa(valor) -> int | None:
    """Convierte el id de empresa (int, float o texto) a int; None si no es válido."""
    if valor is None or valor == "":
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def _clabe_valida(texto) -> bool:
    """True si `texto` contiene una CLABE de 18 dígitos con dígito de control
    correcto (ignora espacios/guiones u otros separadores)."""
    dig = re.sub(r"\D", "", str(texto or ""))
    return len(dig) == 18 and validar_clabe(dig)


def _texto_cuenta(valor) -> str:
    """Cuenta como texto, sin el apóstrofo con que Excel fuerza texto y sin espacios
    en los extremos. Un número entero de Excel se muestra sin el '.0'."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).replace("'", "").strip()


def _leer_excel(ruta: str) -> dict[int, list[dict]]:
    """Lee el Excel. Devuelve {} si no se puede (no existe, bloqueado, formato
    inesperado). Columnas por ENCABEZADO: 'id Empresa' y 'Cuenta' (obligatorias) y,
    opcionales, 'numeroCuenta' y 'CLABE interbancaria'. Cada empresa mapea a una
    lista de registros {'cuenta': str, 'numero': str, 'clabe': str, 'moneda': int|None}
    (sin duplicados); las columnas opcionales ausentes quedan como '' / None."""
    catalogo: dict[int, list[dict]] = {}
    if openpyxl is None or not os.path.exists(ruta):
        return catalogo
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception:
        return catalogo  # p. ej. PermissionError si está abierto en Excel
    try:
        ws = wb[wb.sheetnames[0]]
        filas = ws.iter_rows(values_only=True)
        try:
            encabezados = next(filas)
        except StopIteration:
            return catalogo  # hoja vacía
        idx: dict[str, int] = {}
        for i, celda in enumerate(encabezados or ()):
            clave = _norm(celda)
            if clave and clave not in idx:
                idx[clave] = i
        i_id = next((idx[h] for h in _HDR_ID if h in idx), None)
        i_cta = next((idx[h] for h in _HDR_CUENTA if h in idx), None)
        i_num = next((idx[h] for h in _HDR_NUMERO if h in idx), None)    # opcional
        i_clabe = next((idx[h] for h in _HDR_CLABE if h in idx), None)   # opcional
        i_mon = next((idx[h] for h in _HDR_MONEDA if h in idx), None)    # opcional
        if i_id is None or i_cta is None:
            return catalogo  # no es el Excel esperado -> {}

        def col(fila, j):
            return fila[j] if j is not None and j < len(fila) else None

        for fila in filas:
            if not fila:
                continue
            id_emp = _id_empresa(col(fila, i_id))
            cuenta = _texto_cuenta(col(fila, i_cta))
            numero = _texto_cuenta(col(fila, i_num)) if i_num is not None else ""
            clabe = _texto_cuenta(col(fila, i_clabe)) if i_clabe is not None else ""
            moneda = id_moneda(col(fila, i_mon)) if i_mon is not None else None
            if id_emp is None or not cuenta:
                continue
            registros = catalogo.setdefault(id_emp, [])
            if not any(r["cuenta"] == cuenta and r["clabe"] == clabe
                       for r in registros):  # sin duplicar por empresa
                registros.append({"cuenta": cuenta, "numero": numero,
                                  "clabe": clabe, "moneda": moneda})
    finally:
        wb.close()
    return catalogo


def _cargar(ruta: str) -> dict[int, list[dict]]:
    """Carga el catálogo; si el Excel se puede leer, actualiza el caché; si no,
    usa la última lectura guardada en caché."""
    catalogo = _leer_excel(ruta)
    if catalogo:
        try:  # el JSON exige claves str: se guardan como texto y se reconvierten
            with open(_RUTA_CACHE, "w", encoding="utf-8") as fh:
                json.dump({str(k): v for k, v in catalogo.items()}, fh,
                          ensure_ascii=False)
        except Exception:
            pass
        return catalogo
    if os.path.exists(_RUTA_CACHE):
        try:
            with open(_RUTA_CACHE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            return {int(k): list(v) for k, v in data.items()}
        except Exception:
            pass
    return {}


def instalar_excel(ruta_origen: str) -> int:
    """Instala el Excel elegido en RUTA_EXCEL de forma TRANSACCIONAL y devuelve
    cuántas empresas quedaron con cuentas.

    Respalda el actual, copia el nuevo y lo LEE para validarlo; si no se reconoce,
    hace ROLLBACK y lanza ExcelCuentasDispersionInvalido. Si es válido, invalida el
    caché. La mecánica (incluidos los casos de archivo bloqueado y de elegir el
    archivo desde su propia ruta) vive en `core.instalador_catalogo`."""
    def validar(catalogo):
        if not catalogo:
            raise ExcelCuentasDispersionInvalido(
                "El archivo no tiene el formato esperado (columnas 'id Empresa' y "
                "'Cuenta')."
            )

    catalogo = instalador_catalogo.instalar(
        ruta_origen, RUTA_EXCEL, _leer_excel, validar, _RUTA_CACHE)
    return len(catalogo)


def _de_la_moneda(registro: dict, id_mon: int | None) -> bool:
    """True si el registro sirve para una solicitud de la moneda `id_mon`.

    Sin filtro (`id_mon` None) pasa todo. Una cuenta SIN moneda en el Excel también
    pasa siempre: la columna es opcional y retrocompatible, así que un catálogo
    viejo —o una fila que se quedó sin llenar— no puede dejar una pestaña sin
    ninguna cuenta que elegir y bloquear la dispersión."""
    if id_mon is None:
        return True
    propia = registro.get("moneda")
    return propia is None or propia == id_mon


class CatalogoCuentasDispersion:
    """Acceso al catálogo de cuentas de dispersión por id de empresa."""

    def __init__(self, ruta: str = RUTA_EXCEL):
        self.datos = _cargar(ruta)

    def disponible(self) -> bool:
        return bool(self.datos)

    def empresas(self) -> list[int]:
        return sorted(self.datos.keys())

    def _registros(self, id_empresa) -> list[dict]:
        if id_empresa is None:
            return []
        try:
            clave = int(id_empresa)
        except (TypeError, ValueError):
            return []
        return self.datos.get(clave, [])

    def cuentas_por_id_empresa(self, id_empresa, moneda=None) -> list[str]:
        """Cuentas ('Cuenta') de una empresa por su id. [] si no hay o el id es
        None. Ordenadas (alfabético) y sin duplicar, para el selector.

        `moneda` (id del catálogo o siglas: 'USD', 'MXN'…) acota a las cuentas de esa
        moneda: una solicitud en dólares no se paga desde una cuenta en pesos. Ver
        `_de_la_moneda` para el trato de las cuentas sin moneda."""
        id_mon = id_moneda(moneda)
        cuentas = {r.get("cuenta", "") for r in self._registros(id_empresa)
                   if _de_la_moneda(r, id_mon)}
        return sorted(c for c in cuentas if c)

    def identificadores_de_cuenta(self, id_empresa, cuenta: str) -> list[str]:
        """Todo lo que identifica a `cuenta` en un comprobante bancario, en orden de
        preferencia: **Cuenta > numeroCuenta > CLABE**. Sin vacíos.

        Existe porque el comprobante trae la cuenta enmascarada por sus últimos
        dígitos, y hay cuentas cuyo nombre no tiene ninguno (p. ej. 'PETRO SMART
        HERMOSILLO BBVA'): con solo el nombre, esa regla de casado nunca se cumple.
        Devolviendo los tres candidatos, quien compara puede probarlos todos.

        Si `cuenta` no está en el catálogo devuelve solo el texto recibido, para que
        el llamador siga teniendo algo con que comparar."""
        objetivo = (cuenta or "").strip()
        for r in self._registros(id_empresa):
            if r.get("cuenta", "") == objetivo:
                vals = [objetivo, r.get("numero", ""), r.get("clabe", "")]
                # dict.fromkeys: quita repetidos conservando el orden de preferencia.
                return list(dict.fromkeys(v for v in vals if v))
        return [objetivo] if objetivo else []

    def clabes_por_id_empresa(self, id_empresa, moneda=None) -> list[str]:
        """CLABEs interbancarias VÁLIDAS de una empresa por su id (cuenta origen del
        TXT en pesos). Solo se incluyen las que son una CLABE de 18 dígitos con
        dígito de control correcto. [] si no hay o el id es None. Ordenadas y sin
        duplicar. `moneda` acota por moneda de la cuenta (ver
        `cuentas_por_id_empresa`)."""
        id_mon = id_moneda(moneda)
        clabes = {r.get("clabe", "") for r in self._registros(id_empresa)
                  if _de_la_moneda(r, id_mon)}
        return sorted(c for c in clabes if _clabe_valida(c))

    def cuentas_clabe_por_id_empresa(
        self, id_empresa, moneda=None,
    ) -> list[tuple[str, str]]:
        """Pares (cuenta, clabe) de una empresa donde la CLABE es VÁLIDA y el banco
        de la cuenta TIENE formato de generación en la app (BANREGIO / BBVA /
        BANCOMER). La 'cuenta' es el texto a MOSTRAR (trae banco/empresa, evita
        confusiones) y la CLABE es el valor con el que opera el TXT en pesos. Sin
        duplicar por CLABE y ordenado por el texto de la cuenta. `moneda` acota por
        moneda de la cuenta (ver `cuentas_por_id_empresa`)."""
        id_mon = id_moneda(moneda)
        vistos: set[str] = set()
        pares: list[tuple[str, str]] = []
        for r in self._registros(id_empresa):
            if not _de_la_moneda(r, id_mon):
                continue
            clabe = r.get("clabe", "")
            cuenta = r.get("cuenta", "")
            # Solo bancos con layout soportado (evita cuentas sin formato en la app).
            if not _clabe_valida(clabe) or banco_formato(cuenta) is None \
                    or clabe in vistos:
                continue
            vistos.add(clabe)
            pares.append((cuenta or clabe, clabe))  # fallback: la CLABE como texto
        return sorted(pares, key=lambda p: p[0].casefold())

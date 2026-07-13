"""Catálogo de cuentas bancarias de las empresas que dispersan.

Lee el Excel 'Cuentas bancarias/CUENTAS BANCARIAS .xlsx' (columnas Cuenta,
Alias/Empresa, Divisa, Banco, CLABE y, opcionalmente, 'Alias corto') y permite
consultar las cuentas de una empresa por (empresa, banco) o por 'Alias corto'
(el nombre corto que usa la dispersión, p. ej. "Abastecedora"). Así el usuario
elige y la cuenta aparece sola, sin escribirla.

Las columnas se localizan por ENCABEZADO (fila 1), no por posición fija, para
tolerar reordenamientos y columnas nuevas. Si el Excel se actualiza, basta
reabrir la aplicación para reflejar los cambios.
"""

from __future__ import annotations

import json
import os
import re
import shutil
from dataclasses import asdict, dataclass

try:
    import openpyxl
except ImportError:  # openpyxl es opcional; sin él, el catálogo queda vacío.
    openpyxl = None

from . import rutas

# El Excel lo actualiza el usuario, así que va junto al .exe (no empaquetado).
RUTA_EXCEL = os.path.join(rutas.DATOS, "Cuentas bancarias", "CUENTAS BANCARIAS .xlsx")
# Copia en caché del catálogo. Permite seguir trabajando aunque el Excel esté
# abierto (Excel lo bloquea exclusivamente) usando la última lectura válida.
_RUTA_CACHE = os.path.join(rutas.DATOS, "_cuentas_cache.json")

# Banco elegido en la interfaz -> nombre del banco tal como aparece en el Excel.
BANCO_UI_A_EXCEL = {
    "Banregio": "BANREGIO",
    "Bancomer": "BBVA BANCOMER",
}


@dataclass
class CuentaBancaria:
    """Una cuenta del catálogo, con todos sus datos ya normalizados."""

    clabe: str = ""
    divisa: str = ""            # divisa tal como viene del Excel (PESOS, USD, ...)
    cuenta: str = ""           # número de cuenta ya normalizado (ver _limpiar_cuenta)
    banco: str = ""
    empresa: str = ""          # nombre largo (columna 'Alias')
    alias_corto: str = ""      # nombre corto que usa la dispersión (p. ej. "Abastecedora")
    moneda_general: str = ""   # MXN / USD (ver _moneda_general)


def hay_excel() -> bool:
    """True si ya hay un Excel de cuentas colocado para su consulta."""
    return os.path.exists(RUTA_EXCEL)


class ExcelCuentasInvalido(ValueError):
    """El Excel adjuntado no tiene el formato esperado de cuentas bancarias."""


def instalar_excel(ruta_origen: str) -> int:
    """Instala el Excel elegido en RUTA_EXCEL de forma TRANSACCIONAL y devuelve
    cuántas empresas quedaron cargadas.

    Pasos: respalda el actual (si hay), copia el nuevo, y lo LEE para validarlo.
    Si no se puede leer como catálogo de cuentas (formato inesperado), hace
    ROLLBACK al anterior (o elimina el nuevo si no había) y lanza
    ExcelCuentasInvalido. Si es válido, invalida el caché para que la próxima
    consulta use el archivo nuevo."""
    os.makedirs(os.path.dirname(RUTA_EXCEL), exist_ok=True)
    # Respaldo del actual (si existe) para poder revertir.
    respaldo = None
    if os.path.exists(RUTA_EXCEL):
        respaldo = RUTA_EXCEL + ".bak"
        shutil.copyfile(RUTA_EXCEL, respaldo)
    try:
        shutil.copyfile(ruta_origen, RUTA_EXCEL)
        catalogo = _leer_excel(RUTA_EXCEL)
        if not catalogo:  # no se reconocieron cuentas -> no es el Excel esperado
            raise ExcelCuentasInvalido(
                "El archivo no tiene el formato esperado de cuentas bancarias "
                "(no se encontraron cuentas válidas)."
            )
    except Exception:
        # Rollback: restaura el anterior, o borra el nuevo si no había.
        if respaldo is not None:
            shutil.copyfile(respaldo, RUTA_EXCEL)
        elif os.path.exists(RUTA_EXCEL):
            os.remove(RUTA_EXCEL)
        raise
    finally:
        if respaldo is not None and os.path.exists(respaldo):
            os.remove(respaldo)
    # Éxito: invalida el caché para forzar la relectura del archivo nuevo.
    try:
        if os.path.exists(_RUTA_CACHE):
            os.remove(_RUTA_CACHE)
    except OSError:
        pass  # si no se puede borrar, se sobreescribirá en la próxima lectura
    return len(catalogo)


def _limpiar_clabe(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _limpiar_cuenta(valor) -> str:
    """Normaliza el número de cuenta:
      - quita el apóstrofo con que Excel fuerza texto ("'0510125198" -> "0510125198");
      - si la cuenta es SOLO números (dígitos y espacios): reemplaza los espacios
        internos por '-' y corta los ceros a la izquierda.
    Cuentas con letras se dejan como están (solo se limpia el apóstrofo/espacios extremos)."""
    s = str(valor or "").replace("'", "").strip()
    if s and re.fullmatch(r"[\d\s]+", s):
        s = re.sub(r"^0+", "", s).strip()  # corta TODOS los ceros iniciales
        s = re.sub(r"\s+", "-", s)         # espacio(s) internos -> "-"
        if not s:                          # era todo ceros
            s = "0"
    return s


def _moneda_general(divisa) -> str:
    """Divisa del Excel -> moneda general (MXN / USD). Otros valores (p. ej.
    'CREDITO') se conservan en mayúsculas; '' si viene vacío."""
    d = str(divisa or "").strip().upper()
    if d in ("MXN", "MXP", "PESOS"):
        return "MXN"
    if d in ("DLLS", "USD"):
        return "USD"
    return d


# Tipo del catálogo en memoria: empresa -> banco -> lista de cuentas.
_Catalogo = "dict[str, dict[str, list[CuentaBancaria]]]"

# Encabezados esperados (normalizados) -> campo. 'alias corto' es opcional.
_REQUERIDOS = ("cuenta", "alias", "divisa", "banco", "clabe")


def _norm_encabezado(valor) -> str:
    return str(valor or "").strip().casefold()


def _leer_excel(ruta: str):
    """Lee el Excel. Devuelve {} si no se puede (no existe, bloqueado, formato
    inesperado). Las columnas se localizan por ENCABEZADO en la fila 1:
    Cuenta, Alias (empresa), Divisa, Banco, CLABE y, opcional, 'Alias corto'."""
    catalogo: dict = {}
    if openpyxl is None or not os.path.exists(ruta):
        return catalogo
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception:
        return catalogo  # p. ej. PermissionError si está abierto en Excel
    try:
        nombre_hoja = (
            "RESUMEN DE CUENTAS" if "RESUMEN DE CUENTAS" in wb.sheetnames
            else wb.sheetnames[0]
        )
        ws = wb[nombre_hoja]
        filas = ws.iter_rows(values_only=True)
        try:
            encabezados = next(filas)
        except StopIteration:
            return catalogo  # hoja vacía
        idx = {}
        for i, celda in enumerate(encabezados or ()):
            clave = _norm_encabezado(celda)
            if clave and clave not in idx:
                idx[clave] = i
        if not all(h in idx for h in _REQUERIDOS):
            return catalogo  # no es el Excel de cuentas esperado -> {}
        i_corto = idx.get("alias corto")

        def col(fila, j):
            return fila[j] if j is not None and j < len(fila) else None

        for fila in filas:
            if not fila:
                continue
            empresa = str(col(fila, idx["alias"]) or "").strip()
            banco = str(col(fila, idx["banco"]) or "").strip()
            clabe = _limpiar_clabe(col(fila, idx["clabe"]))
            if not empresa or not banco or len(clabe) != 18:
                continue
            divisa = str(col(fila, idx["divisa"]) or "").strip()
            reg = CuentaBancaria(
                clabe=clabe,
                divisa=divisa,
                cuenta=_limpiar_cuenta(col(fila, idx["cuenta"])),
                banco=banco,
                empresa=empresa,
                alias_corto=str(col(fila, i_corto) or "").strip(),
                moneda_general=_moneda_general(divisa),
            )
            catalogo.setdefault(empresa, {}).setdefault(banco, []).append(reg)
    finally:
        wb.close()
    return catalogo


def _serializar(catalogo) -> dict:
    return {
        empresa: {banco: [asdict(r) for r in regs] for banco, regs in bancos.items()}
        for empresa, bancos in catalogo.items()
    }


def _deserializar(data) -> dict:
    """Reconstruye el catálogo desde el caché JSON. Tolera el formato viejo
    (listas [clabe, divisa, cuenta]) rellenando los campos nuevos."""
    catalogo: dict = {}
    for empresa, bancos in (data or {}).items():
        for banco, regs in (bancos or {}).items():
            for r in regs:
                if isinstance(r, dict):
                    cta = CuentaBancaria(**{
                        k: r.get(k, "") for k in CuentaBancaria().__dict__
                    })
                else:  # formato viejo: [clabe, divisa, cuenta]
                    clabe, divisa, cuenta = (list(r) + ["", "", ""])[:3]
                    cta = CuentaBancaria(
                        clabe=clabe, divisa=divisa, cuenta=cuenta, banco=banco,
                        empresa=empresa, alias_corto="",
                        moneda_general=_moneda_general(divisa),
                    )
                catalogo.setdefault(empresa, {}).setdefault(banco, []).append(cta)
    return catalogo


def _cargar(ruta: str):
    """Carga el catálogo; si el Excel se puede leer, actualiza el caché; si no
    (p. ej. está abierto en Excel), usa la última lectura guardada en caché."""
    catalogo = _leer_excel(ruta)
    if catalogo:
        try:  # guarda la última versión buena
            with open(_RUTA_CACHE, "w", encoding="utf-8") as fh:
                json.dump(_serializar(catalogo), fh, ensure_ascii=False)
        except Exception:
            pass
        return catalogo
    # No se pudo leer el Excel -> intentar el caché.
    if os.path.exists(_RUTA_CACHE):
        try:
            with open(_RUTA_CACHE, "r", encoding="utf-8") as fh:
                return _deserializar(json.load(fh))
        except Exception:
            pass
    return {}


class CatalogoCuentas:
    """Acceso al catálogo de cuentas por empresa y banco (o por 'Alias corto')."""

    def __init__(self, ruta: str = RUTA_EXCEL):
        self.datos = _cargar(ruta)
        # Índice por 'Alias corto' (normalizado) construido una sola vez, para
        # que la dispersión resuelva las cuentas de una empresa en O(1).
        self._por_alias: dict[str, list[CuentaBancaria]] = {}
        for bancos in self.datos.values():
            for regs in bancos.values():
                for r in regs:
                    if r.alias_corto:
                        clave = r.alias_corto.strip().casefold()
                        self._por_alias.setdefault(clave, []).append(r)

    def disponible(self) -> bool:
        return bool(self.datos)

    def empresas(self) -> list[str]:
        return sorted(self.datos.keys())

    def cuentas(self, empresa: str, banco_ui: str) -> list[tuple[str, str, str]]:
        """Cuentas (clabe, divisa, num_cuenta) de una empresa para el banco
        elegido en la UI. Las cuentas en PESOS/MXP se listan primero."""
        banco = BANCO_UI_A_EXCEL.get(banco_ui, banco_ui)
        regs = list(self.datos.get(empresa, {}).get(banco, []))
        regs.sort(key=lambda r: 0 if r.divisa.upper() in ("PESOS", "MXP") else 1)
        return [(r.clabe, r.divisa, r.cuenta) for r in regs]

    def cuentas_por_alias_corto(self, alias_corto: str) -> list[CuentaBancaria]:
        """Cuentas cuyo 'Alias corto' coincide con `alias_corto` (ignorando
        mayúsculas/espacios). Orden: alfabético por banco; las de dólares (USD)
        se listan al final."""
        regs = list(self._por_alias.get((alias_corto or "").strip().casefold(), []))
        regs.sort(key=lambda r: (r.moneda_general == "USD", r.banco.casefold(), r.cuenta))
        return regs

"""Genera el libro de saldos: el mismo formato de tesorería, con datos frescos.

No dibuja un reporte. Abre el libro base —el formato real, vaciado, con sus 24
pestañas y sus 746 fórmulas intactas— y pega cada descarga bancaria en su
pestaña, EN LA FILA QUE LE TOCA A ESA CUENTA. Es el copiar-y-pegar que hoy hace
el macro a mano, pero colocando por número de cuenta en vez de por el orden en
que el portal quiso entregar las filas.

De ahí sale la propiedad que hace que valga la pena: **las fórmulas de SALDOS no
se tocan y quedan correctas por construcción**. `=HSBC!C2` lee la fila 2 de HSBC,
y la fila 2 de HSBC es la cuenta que debe ser porque nosotros la pusimos ahí. Lo
mismo vale para SALDOS HORIZOTAL, que sale bien sola sin escribir una línea para
ella.

Por eso aquí se escriben VALORES solo en las pestañas de descarga y en los
ledgers de flujo. Ni un total, ni un subtotal, ni el desglose Combustibles/Resto:
todo eso ya está en el base como fórmula y se recalcula al abrir en Excel.

## Lo que no se inserta

Una cuenta que no está en la plantilla NO se agrega a su pestaña. Insertar una
fila correría todo lo de abajo y rompería cada referencia que apunte ahí — que es
exactamente el defecto que este módulo existe para eliminar. Esas cuentas van a
la pestaña "Cuentas nuevas", con su monto, para que se vean y se decida si el
formato debe crecer. Las cuentas de CUENTAS_EXCLUIDAS (core/saldos.py) no llegan
ni ahí: se descartan antes y el libro no deja rastro de ellas, que es justo lo
que se pide de una exclusión.

## openpyxl no calcula

Las celdas de fórmula salen sin valor cacheado, así que el archivo recién
generado se ve vacío en cualquier lector que no evalúe (incluido openpyxl). Los
números aparecen al abrirlo en Excel. Es esperado: el cuadre se verifica ahí.
"""

from __future__ import annotations

import datetime
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from .saldos_plantilla import cargar as cargar_plantilla

# La hoja que se agrega al final con lo que no cupo en el formato.
HOJA_EXCEPCIONES = "Cuentas nuevas"

# Rol especial: la celda de la fila SIGUIENTE donde BAJÍO guarda el número de
# cuenta ('Cuenta: 16084470201Conecta BanBajio').
_ROL_CUENTA_ABAJO = "_fila_cuenta_abajo"

# Dónde va el sello de actualización en la hoja SALDOS. Son las celdas que el
# formato ya usaba para eso (`K3`/`K5` traen las etiquetas 'Fecha:' y 'Hora:').
# La cabecera lleva DOS pares fecha/hora: L3-L4 son los de hoy y L5-L6 los del
# día hábil anterior. La hora de hoy va en L4, no en L5: ponerla en L5 pisaba la
# fecha de la comparativa.
CELDA_FECHA = "L3"
CELDA_HORA = "L4"

# El calendario de flujo lleva la fecha de cada día en la columna L, pero SOLO el
# lunes está escrito: los demás son `=+L10+1` y encadenan solos. Son cuatro
# anclas, una por panel, y de ellas cuelgan los paneles de abajo (`L50=+L30`,
# `L60=+L40`, `L68=+L60`), así que basta con escribir estas cuatro para que las
# siete semanas queden en su sitio.
#
# Estaban clavadas en el formato —traían la semana del 24 de agosto— y nadie las
# movía al generar: el reporte salía con el calendario de una semana vieja y los
# importes de esta, que es peor que no traer fecha.
CELDAS_SEMANA = ("L10", "L20", "L30", "L40")

# Lo que tesorería CAPTURA A MANO en el calendario de flujo y hay que conservar
# durante la semana. El formato no lo calcula: son celdas vacías que se llenan en
# Excel, así que cada reporte nuevo salía en blanco y había que recapturarlas.
#
# La columna la fija el propio formato en la fila 8: M='Pago', N='Saldo',
# O='Importe'. Los ocho paneles son TOTAL (10-16), PEMEX (20), MGC (30), TESORO
# (40), ACP (50), NÓMINA (60), IMPUESTOS (68) y CRÉDITOS (76).
#
# El panel TOTAL queda FUERA a propósito: su `M10` es `=+M20+M30+M50+…`, la suma
# de los demás. Escribir ahí un valor destruiría la consolidación.
_PANELES_MANUALES = ((20, 26), (30, 36), (40, 46), (50, 56), (60, 64),
                     (68, 72), (76, 82))
CELDAS_PAGOS = tuple("M{}".format(f)
                     for ini, fin in _PANELES_MANUALES
                     for f in range(ini, fin + 1))
# El importe solo se teclea en los dos paneles que no lo derivan de un ledger:
# ACP e IMPUESTOS. En los demás es fórmula y no se toca.
CELDAS_IMPORTES = tuple("O{}".format(f)
                        for ini, fin in ((50, 56), (68, 72))
                        for f in range(ini, fin + 1))
CELDAS_MANUALES = CELDAS_PAGOS + CELDAS_IMPORTES
# Las filas que abarcan, para leerlas de un tirón sin abrir el libro entero.
_FILAS_MANUALES = frozenset(int(c[1:]) for c in CELDAS_MANUALES)

_FMT_FECHA = "dd/mm/yyyy"
_FMT_HORA = "hh:mm"
# Dos decimales, negativos en rojo entre paréntesis. El formato original recortaba
# a enteros, lo que hacía que el reporte y el resumen en pantalla no cuadraran a
# la vista aunque las celdas fueran correctas.
_FMT_SALDO = "#,##0.00_);[Red](#,##0.00)"

_ENC_EXCEPCIONES = (
    ("Tipo", 14),
    ("Banco", 22),
    ("Cuenta", 24),
    ("Titular en el reporte", 44),
    ("Saldo", 18),
    ("Moneda", 10),
    ("Archivo de origen", 40),
    ("Motivo", 46),
)
_AZUL = "FF317FB1"


class ErrorExport(Exception):
    """No se pudo generar el libro."""


def _col(hoja, letra, fila):
    return hoja.cell(fila, column_index_from_string(letra))


def _numero_de_cuenta(cols, destino, linea):
    """El número tal como lo escribe esa pestaña del formato.

    Cuando la fila ya venía numerada en el formato se respeta esa forma. Cuando
    no —las que se resuelven por los últimos 4 dígitos— hay que elegirla: si la
    región tiene columna de sucursal aparte, va la cuenta CORTA, porque es lo que
    llevan sus vecinas y mezclar las dos formas en la misma columna vuelve la
    pestaña ilegible."""
    if destino.cuenta:
        return str(destino.cuenta)
    if "sucursal" in cols:
        corta = (linea.extra or {}).get("cuenta_corta")
        if corta:
            return str(corta)
    return str(linea.cuenta)


def _escribir_linea(hoja, fila, cols, destino, linea, moneda_formato):
    """Vuelca el SALDO de una línea leída en su fila canónica.

    La identidad de la fila —número, titular, sucursal…— ya la puso
    `_escribir_identidad` con los datos del formato; aquí solo va lo que cambia
    cada día. La excepción son las filas que el formato NO numeraba: ahí el
    número, la sucursal y el titular salen del propio archivo, porque no hay de
    dónde más sacarlos."""
    if not destino.cuenta:
        if "cuenta" in cols:
            celda = _col(hoja, cols["cuenta"], fila)
            celda.value = _numero_de_cuenta(cols, destino, linea)
            celda.number_format = "@"
        if "sucursal" in cols:
            # Banamex reporta sucursal y cuenta en columnas separadas y el lector
            # las concatena. Al escribir hay que volver a partirlas.
            corta = str((linea.extra or {}).get("cuenta_corta") or "")
            completa = str(linea.cuenta or "")
            if corta and completa.endswith(corta):
                _col(hoja, cols["sucursal"], fila).value = (
                    completa[:len(completa) - len(corta)] or None)
        if "titular" in cols and not destino.titular:
            _col(hoja, cols["titular"], fila).value = linea.titular or ""
        if "moneda" in cols:
            _col(hoja, cols["moneda"], fila).value = moneda_formato

    # El saldo va a la columna de la región y, si algún renglón de SALDOS lee otra
    # (Banorte tiene uno que toma 'Saldo actual' en vez de 'Disponible'), también
    # a esa. Solo a esas dos: rellenar columnas que no alimentan nada sería
    # inventar un dato que el portal no dio.
    columnas = set()
    if "saldo" in cols:
        columnas.add(cols["saldo"])
    if destino.renglon is not None:
        columnas.add(destino.renglon.hoja_col)
    for letra in columnas:
        celda = _col(hoja, letra, fila)
        celda.value = linea.saldo
        celda.number_format = _FMT_SALDO


def _escribir_identidad(libro, plantilla):
    """Repone en cada pestaña lo que identifica a sus cuentas, fila por fila.

    Va ANTES de los saldos y cubre TODAS las filas del mapa, hayan reportado o
    no. Es lo que hace que una cuenta que el portal no trajo hoy siga apareciendo
    en su pestaña —con su número, su titular y su sucursal— y solo el saldo en
    blanco. Antes desaparecía entera: a la vista faltaba una cuenta y nada lo
    decía.

    Son datos que NO cambian de un día a otro, así que se conservan del formato y
    no del archivo. El archivo solo aporta el saldo."""
    for nombre, info in plantilla.hojas.items():
        hoja = libro[nombre]
        for fila, entrada in info["filas"].items():
            fila = int(fila)
            cols = plantilla.columnas(nombre, fila)
            for rol, valor in (entrada.get("estaticos") or {}).items():
                if rol == _ROL_CUENTA_ABAJO:
                    # BAJÍO guarda el número en la fila de ABAJO, en la columna
                    # del titular.
                    letra, destino_fila = cols.get("titular"), fila + 1
                else:
                    letra, destino_fila = cols.get(rol), fila
                if not letra:
                    continue
                celda = _col(hoja, letra, destino_fila)
                celda.value = valor
                if rol in ("cuenta", "clabe", "sucursal"):
                    celda.number_format = "@"   # texto: conserva ceros al inicio


def _moneda_para(destino, linea):
    """Divisa en el vocabulario de esa pestaña.

    Cada portal la nombra distinto — BBVA y Banorte dicen 'MXP', Banregio 'MXN',
    Intercam 'PESOS'. El base conserva la que traía el formato para esa fila; se
    respeta, y solo se recurre a la del lector cuando no hay."""
    if destino.moneda:
        return destino.moneda
    return (linea.moneda or "MXN").upper()


def _escribir_saldos(libro, asignacion):
    """Pega todas las líneas colocadas en sus pestañas.

    Una cuenta puede ocupar más de una fila —el formato repite algunas en su
    bloque de "cuentas nuevas"— y el saldo va a todas, igual que cuando se pega a
    mano."""
    plantilla = asignacion.plantilla
    for colocada in asignacion.colocadas.values():
        destino = colocada.destino
        for d in (destino,) + tuple(destino.gemelos):
            cols = plantilla.columnas(d.hoja, d.fila)
            if not cols:
                continue
            _escribir_linea(libro[d.hoja], d.fila, cols, d, colocada.linea,
                            _moneda_para(d, colocada.linea))


def _escribir_ledgers(libro, plantilla, insumos):
    """Vuelca los insumos de flujo en sus pestañas.

    Los `SUMIF` de SALDOS ya apuntan a estos rangos: basta con poner cada dato en
    la columna que le toca. De los ledgers de vencimientos se escriben solo las
    columnas que alguna fórmula lee (fecha e importe) más la referencia, que no
    alimenta nada pero permite auditar de dónde salió una cifra.

    Un insumo que no vino no se escribe y su panel queda en cero. Que falte el
    archivo de nómina no puede impedir que salga el reporte de saldos."""
    escritos = {}
    for nombre, datos in (insumos or {}).items():
        info = plantilla.ledgers.get(nombre)
        if info is None or not datos:
            continue
        hoja = _hoja_por_clave(libro, nombre)
        if info.get("modo") == "copia":
            escritos[nombre] = _copiar_rangos(hoja, datos)
            continue
        escritos[nombre] = _escribir_vencimientos(hoja, info, datos)
    return escritos


def _escribir_vencimientos(hoja, info, filas):
    """Vuelca un ledger de vencimientos en su pestaña.

    Va la fila COMPLETA, no solo las columnas que alimentan los `SUMIF`: el
    reporte se coteja contra el sistema de origen, y para eso hacen falta el
    número de documento, la clase, las fechas y el texto de cabecera. Encima se
    reescriben fecha e importe ya interpretados, con su formato, porque los
    `SUMIF` comparan contra fechas reales."""
    cols = info.get("cols") or {}
    if not cols:
        return 0
    tope = info["fila_fin"]
    fila = info["fila_ini"]
    for registro in filas:
        if fila > tope:
            break
        for i, valor in enumerate(registro.get("celdas") or (), start=1):
            if valor is not None:
                hoja.cell(fila, i).value = valor
        # Solo fecha e importe: son los que los `SUMIF` necesitan interpretados.
        # La referencia ya viaja en `celdas`.
        for rol in ("fecha", "importe"):
            letra, valor = cols.get(rol), registro.get(rol)
            if not letra or valor is None:
                continue
            celda = _col(hoja, letra, fila)
            celda.value = valor
            celda.number_format = (_FMT_FECHA if rol == "fecha" else _FMT_SALDO)
        fila += 1
    escritas = fila - info["fila_ini"]
    if escritas < len(filas):
        # Truncar en silencio dejaría un total que parece bueno y no lo es.
        raise ErrorExport(
            "el insumo trae {} filas y en la hoja caben {}; hay que ampliar el "
            "rango en saldos_mapa.json".format(len(filas), escritas))
    return escritas


def _copiar_rangos(hoja, datos):
    """Copia bloques de celdas a sus mismas coordenadas (caso CRÉDITOS)."""
    total = 0
    for rango in datos.get("rangos", ()):
        col_ini = column_index_from_string(rango["col_ini"])
        for i, valores in enumerate(rango["celdas"]):
            fila = rango["fila_ini"] + i
            for j, valor in enumerate(valores):
                if valor is None:
                    continue
                hoja.cell(fila, col_ini + j).value = valor
                total += 1
    return total


def _hoja_por_clave(libro, clave):
    """Busca la hoja por su nombre sin acentos ('CREDITOS' -> 'CRÉDITOS')."""
    import unicodedata
    objetivo = clave.upper().strip()
    for nombre in libro.sheetnames:
        plano = unicodedata.normalize("NFKD", nombre)
        plano = "".join(c for c in plano if not unicodedata.combining(c))
        if plano.upper().strip() == objetivo:
            return libro[nombre]
    raise ErrorExport("el libro base no tiene la hoja {!r}".format(clave))


def _hoja_excepciones(libro, asignacion):
    """Pestaña con lo que llegó y no cupo en el formato.

    Separa cuentas NUEVAS de DUPLICADAS en una columna, cosa que la versión
    anterior no hacía: las mezclaba y no había forma de distinguir un hallazgo
    (se abrió una cuenta) de un aviso (el mismo archivo se subió dos veces).

    Las excluidas NO salen aquí: se descartan en la identificación y el libro no
    las menciona en ninguna parte."""
    filas = ([("Nueva", s) for s in asignacion.nuevas]
             + [("Duplicada", s) for s in asignacion.duplicados])
    if not filas:
        return
    hoja = libro.create_sheet(HOJA_EXCEPCIONES)
    hoja.sheet_state = "visible"

    hoja["A1"] = "Cuentas que llegaron y no están en el formato"
    hoja["A1"].font = Font(name="Calibri", size=14, bold=True)
    hoja["A2"] = ("Para que entren al reporte hay que regenerar la plantilla; "
                  "insertarlas a mano correría las filas y rompería las fórmulas.")
    hoja["A2"].font = Font(name="Calibri", size=10, italic=True)

    for i, (titulo, ancho) in enumerate(_ENC_EXCEPCIONES, start=1):
        celda = hoja.cell(3, i)
        celda.value = titulo
        celda.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        hoja.column_dimensions[get_column_letter(i)].width = ancho

    for i, (tipo, suelta) in enumerate(filas, start=4):
        linea = suelta.linea
        hoja.cell(i, 1).value = tipo
        hoja.cell(i, 2).value = linea.banco
        celda = hoja.cell(i, 3)
        celda.value = str(linea.cuenta)
        celda.number_format = "@"
        hoja.cell(i, 4).value = linea.titular or ""
        celda = hoja.cell(i, 5)
        celda.value = linea.saldo
        celda.number_format = _FMT_SALDO
        hoja.cell(i, 6).value = linea.moneda
        hoja.cell(i, 7).value = linea.origen or ""
        hoja.cell(i, 8).value = suelta.motivo

    hoja.freeze_panes = "A4"


def calcular_totales_cabecera(plantilla, asignacion) -> dict:
    """Los totales de la cabecera de HOY, calculados en Python.

    Hacen falta porque openpyxl no evalúa fórmulas: al generar no se puede LEER
    lo que dará `Q3`, y ese número hay que guardarlo para que mañana caiga en la
    fila del día hábil anterior.

    El mapa ya trae cada total aplanado a los renglones que suma —el desglose
    Combustibles/Resto incluido—, así que aquí solo se suman los saldos que se
    colocaron. Un renglón sin saldo cuenta como cero, igual que una celda vacía
    en Excel.

    OJO con el criterio del formato: `S3 + U3` NO da `Q3`. Tres cuentas
    (Fundación Seren y dos de Operaciones Temáticas) entran al total de pesos y a
    ningún desglose. Se replica tal cual: es su regla de negocio, no un error
    nuestro que toque arreglar por dentro."""
    bandas = {n: b["saldo"] for n, b in plantilla.bandas.items()}
    valor_por_celda = {}
    for colocada in asignacion.colocadas.values():
        renglon = colocada.renglon
        if renglon is None:
            continue
        celda = "{}{}".format(bandas[renglon.banda], renglon.fila)
        valor_por_celda[celda] = colocada.saldo or 0.0

    return {celda: round(sum(valor_por_celda.get(c, 0.0) for c in celdas), 2)
            for celda, celdas in (plantilla.totales_cabecera or {}).items()}


def _escribir_dia_anterior(libro, plantilla, anterior):
    """Copia los totales de la corrida anterior a las filas del día hábil previo.

    En el formato manual esas celdas se llenan pegando a mano los totales del
    reporte de ayer. Aquí se toman del histórico, que guarda una entrada por
    fecha: se usa la más reciente ANTERIOR a hoy, así regenerar el reporte el
    mismo día no borra la comparativa contra ayer."""
    if not anterior:
        return 0
    fecha, hora, totales = anterior
    hoja = libro["SALDOS"]
    escritas = 0
    for origen, destino in (plantilla.espejo_totales or {}).items():
        if origen not in totales:
            continue
        celda = hoja[destino]
        celda.value = totales[origen]
        celda.number_format = _FMT_SALDO
        escritas += 1
    if fecha:
        celda = hoja[plantilla.celda_fecha_anterior]
        celda.value = datetime.datetime.fromisoformat(fecha)
        celda.number_format = _FMT_FECHA
    if hora:
        try:
            celda = hoja[plantilla.celda_hora_anterior]
            celda.value = datetime.datetime.fromisoformat(fecha + "T" + hora)
            celda.number_format = _FMT_HORA
        except ValueError:
            pass
    return escritas


def _sellar(libro, fecha):
    """Pone fecha y hora de actualización, y la semana del calendario de flujo.

    La semana se deriva de la MISMA `fecha` con la que se sella el reporte, no de
    `date.today()`: si algún día se regenera un reporte con fecha de ayer, el
    calendario tiene que ser el de ayer y no el de hoy."""
    hoja = libro["SALDOS"]
    celda = hoja[CELDA_FECHA]
    celda.value = datetime.datetime(fecha.year, fecha.month, fecha.day)
    celda.number_format = _FMT_FECHA
    celda = hoja[CELDA_HORA]
    celda.value = fecha
    celda.number_format = _FMT_HORA

    # Lunes de la semana en curso. Se escribe solo en las anclas; el resto de la
    # columna son fórmulas del formato y NO se tocan, para que Excel siga
    # calculando los días como siempre lo hizo.
    #
    # No se le pone `number_format`: las celdas ya traen el suyo del libro base
    # (`dd/mm/yy;@`) y pisarlo cambiaría cómo se ve el formato sin que nadie lo
    # haya pedido.
    lunes = _lunes_de(fecha)
    for referencia in CELDAS_SEMANA:
        hoja[referencia].value = lunes


def _lunes_de(fecha) -> datetime.datetime:
    """El lunes de la semana que contiene `fecha`, a medianoche.

    `weekday()` da 0 el lunes, así que restarlo cae siempre en el lunes de esa
    misma semana, incluso en fin de semana: un reporte del sábado sigue mostrando
    la semana que empezó el lunes anterior, que es la que está corriendo."""
    dia = datetime.date(fecha.year, fecha.month, fecha.day)
    dia -= datetime.timedelta(days=dia.weekday())
    return datetime.datetime(dia.year, dia.month, dia.day)


def leer_manuales(ruta: str) -> dict:
    """Rescata de un reporte YA GENERADO lo que se capturó a mano en él.

    Es la única forma de conservarlo: esos valores se teclean en Excel, la app
    nunca los ve. Al generar el reporte del día siguiente se vuelven a poner.

    Se ignoran las celdas vacías y las que traigan fórmula —si alguien pegó una,
    la fórmula del formato manda—. Nunca lanza: no poder recuperar la captura de
    ayer no puede impedir el reporte de hoy."""
    if not ruta or not os.path.exists(ruta):
        return {}
    # `read_only` y solo las filas que interesan. Abrir el libro completo para
    # leer 57 celdas costaba DIEZ SEGUNDOS —hay que parsear las 24 pestañas, con
    # los 32 000 renglones de los ledgers— y eso ocurría antes de que apareciera
    # la pantalla de espera: el usuario veía la ventana quieta y volvía a pulsar
    # «Generar». En modo perezoso se corta en la fila 82 y baja a una décima.
    try:
        libro = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001 — un reporte ilegible se trata como ausente
        return {}
    try:
        try:
            hoja = libro["SALDOS"]
        except KeyError:
            return {}
        primera = min(_FILAS_MANUALES)
        ultima = max(_FILAS_MANUALES)
        filas = {primera + i: f for i, f in enumerate(
            hoja.iter_rows(min_row=primera, max_row=ultima, values_only=True))}
        fuera = {}
        for celda in CELDAS_MANUALES:
            columna, fila = celda[0], int(celda[1:])
            valores = filas.get(fila)
            if not valores:
                continue
            i = column_index_from_string(columna) - 1
            valor = valores[i] if i < len(valores) else None
            # Se ignoran las vacías y las que traigan fórmula: si alguien pegó
            # una, la del formato manda.
            if valor is None or isinstance(valor, str):
                continue
            fuera[celda] = valor
        return fuera
    finally:
        libro.close()


def _escribir_manuales(libro, manuales: dict) -> int:
    """Vuelve a poner lo capturado a mano. Devuelve cuántas celdas se llenaron."""
    if not manuales:
        return 0
    hoja = libro["SALDOS"]
    puestas = 0
    for celda, valor in manuales.items():
        # Solo las celdas declaradas: un JSON viejo o tocado a mano no puede
        # escribir en cualquier parte de la hoja.
        if celda not in CELDAS_MANUALES or valor is None:
            continue
        hoja[celda].value = valor
        puestas += 1
    return puestas


def generar(ruta: str, asignacion, insumos: dict = None,
            fecha: datetime.datetime = None, anterior: tuple = None,
            manuales: dict = None) -> dict:
    """Escribe el libro completo en `ruta` y devuelve qué tanto se pudo llenar.

    `asignacion` es lo que devuelve `saldos.identificar()`. `insumos` es opcional:
    {'CREDITOS': [...], 'PEMEX': [...], ...} con las filas de cada ledger.
    `anterior` es `(fecha_iso, hora, {celda: total})` de la corrida previa, que se
    escribe en las filas del día hábil anterior de la cabecera.

    Entre lo que devuelve va `totales_cabecera`: los totales de HOY calculados,
    para que quien llame los guarde y mañana se los pase como `anterior`."""
    plantilla = asignacion.plantilla or cargar_plantilla()
    fecha = fecha or datetime.datetime.now()

    if not os.path.exists(plantilla.ruta_base):
        raise ErrorExport(
            "falta el libro base {}. Genéralo con:\n"
            "  python scripts/derivar_plantilla_saldos.py <formato.xlsx>"
            .format(plantilla.ruta_base))

    libro = openpyxl.load_workbook(plantilla.ruta_base)
    _escribir_identidad(libro, plantilla)
    _escribir_saldos(libro, asignacion)
    ledgers = _escribir_ledgers(libro, plantilla, insumos)
    _hoja_excepciones(libro, asignacion)
    totales_cabecera = calcular_totales_cabecera(plantilla, asignacion)
    comparativas = _escribir_dia_anterior(libro, plantilla, anterior)
    capturadas = _escribir_manuales(libro, manuales)
    _sellar(libro, fecha)
    libro.save(ruta)

    vacios = asignacion.vacios
    return {
        "renglones": asignacion.total_renglones,
        "llenos": asignacion.llenos,
        "pegadas": asignacion.pegadas,
        "vacios": len(vacios),
        "nuevas": len(asignacion.nuevas),
        "duplicados": len(asignacion.duplicados),
        "bancos_faltantes": asignacion.bancos_faltantes(),
        "totales": asignacion.totales(),
        "ledgers": ledgers,
        "ledgers_faltantes": sorted(set(plantilla.ledgers) - set(ledgers)),
        "totales_cabecera": totales_cabecera,
        "comparativas": comparativas,
        "capturadas": capturadas,
    }

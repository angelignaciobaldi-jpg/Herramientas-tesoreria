"""Generación del reporte Excel de la dispersión de devoluciones."""

from __future__ import annotations

import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_AZUL = "1F4E78"
_GRIS = "D9D9D9"
_BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def _fmt_fecha(fecha: str) -> str:
    """Convierte DDMMAAAA -> dd/mm/aaaa. Si no son 8 dígitos, la deja igual."""
    s = re.sub(r"\D", "", fecha or "")
    return f"{s[0:2]}/{s[2:4]}/{s[4:8]}" if len(s) == 8 else (fecha or "")


def generar(ruta: str, registros: list[dict]) -> None:
    """Crea el archivo Excel.

    Cada registro lleva SU PROPIA cuenta origen (empresa/banco/cuenta/número),
    porque en un mismo reporte pueden convivir movimientos asignados a distintas
    cuentas de pago; no se usa un contexto único para todos.

    Args:
        ruta: ruta destino .xlsx
        registros: lista de dicts con las claves:
            empresa, banco, cuenta_origen, num_cuenta, fecha (origen del pago) y
            clabe, monto, beneficiario, concepto (datos del movimiento).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devoluciones"

    # --- Título ---
    ws["A1"] = "Reporte de dispersión de devoluciones"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)

    # --- Tabla (los datos de empresa/banco/cuenta van como columnas, para poder
    #     acumular dispersiones de distintas empresas/bancos en un mismo reporte
    #     maestro) ---
    fila = 3
    encabezados = ["#", "Empresa que paga", "Banco origen", "Cuenta origen (CLABE)",
                   "Número de cuenta", "CLABE Beneficiario", "Monto", "Beneficiario",
                   "Concepto / Referencia", "Fecha de devolución"]
    for col, titulo in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDE

    # Columnas centradas (#, cuentas, CLABE, fecha) y la de monto (con formato).
    CENTRADAS = {1, 4, 5, 6, 10}
    COL_MONTO = 7
    fila_inicio = fila + 1
    for i, reg in enumerate(registros, start=1):
        valores = [i, reg.get("empresa", ""), reg.get("banco", ""),
                   reg.get("cuenta_origen", ""), reg.get("num_cuenta", ""),
                   reg.get("clabe", ""), float(reg.get("monto") or 0),
                   reg.get("beneficiario", ""), reg.get("concepto", ""),
                   _fmt_fecha(reg.get("fecha", ""))]
        for col, valor in enumerate(valores, start=1):
            c = ws.cell(row=fila_inicio + i - 1, column=col, value=valor)
            c.border = _BORDE
            if col in CENTRADAS:
                c.alignment = Alignment(horizontal="center")
            if col == COL_MONTO:
                c.number_format = '#,##0.00'

    # --- Total de montos ---
    fila_total = fila_inicio + len(registros)
    ws.cell(row=fila_total, column=COL_MONTO - 1, value="TOTAL").font = Font(bold=True)
    ct = ws.cell(row=fila_total, column=COL_MONTO,
                 value=f"=SUM(G{fila_inicio}:G{fila_total - 1})")
    ct.font = Font(bold=True)
    ct.number_format = '#,##0.00'
    ct.fill = PatternFill("solid", fgColor=_GRIS)

    # --- Anchos de columna ---
    anchos = {"A": 5, "B": 38, "C": 14, "D": 22, "E": 18, "F": 22,
              "G": 16, "H": 30, "I": 28, "J": 18}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(ruta)

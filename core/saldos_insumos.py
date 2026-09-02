"""Insumos de flujo: lo que alimenta los paneles de créditos y proyección.

La hoja SALDOS no solo lleva saldos. En sus columnas C-H va la posición de
créditos y en K-O un calendario semanal que proyecta el efectivo día por día.
Esos números NO salen de ningún portal bancario: vienen de cinco fuentes que
tesorería ya maneja aparte.

Cuatro son ledgers de vencimientos y SALDOS los consulta siempre igual:

    SUMIF(<columna de fecha>, día, <columna de importe>)

  PEMEX    cuentas por pagar a Pemex        (fecha de vencimiento, saldo)
  MGC      cartera del ERP                  (vencimiento neto, importe)
  TESORO   facturas por pagar               (fecha vencimiento, monto)
  NOMINA   calendario de nómina             (fecha de pago, importe)

La quinta, CRÉDITOS, es distinta: no es una descarga sino la hoja que tesorería
mantiene a mano, con las líneas de crédito por cuenta y la matriz de
amortizaciones por día. SALDOS la lee por posición (`=+CRÉDITOS!F9`), así que se
copia tal cual, en las mismas coordenadas. Eso sería inaceptable en una pestaña
de banco —es justo el defecto que este módulo vino a eliminar— pero aquí quien
fija las filas es una persona que las mantiene estables, no el orden en que un
portal decidió exportar ese día.

Todos los insumos son OPCIONALES. Si no se sube el de nómina, su panel queda en
cero y el reporte de saldos sale completo igual. Nunca bloquean.
"""

from __future__ import annotations

import datetime
import os

from .saldos_lectores import (ErrorLector, _buscar_encabezado, _filas_xls,
                              _filas_xlsx, _a_float, _norm)

# Nombre del insumo -> (alias de columnas, columnas obligatorias). Los alias
# están normalizados igual que en saldos_lectores: minúsculas y sin acentos.
#
# La columna de REFERENCIA es obligatoria aunque ninguna fórmula la use, y es a
# propósito: `_indice_encabezado` casa por igualdad O por prefijo, así que un
# alias como "saldo" se traga «Saldo Retenido» y «Saldo Disponible» y hacía que
# el reporte de BAJÍO se detectara como el ledger de Pemex. Exigir además la
# columna que sí es distintiva de cada insumo («Número de documento»,
# «Referencia», «Factura») cierra ese hueco.
_ALIAS = {
    "PEMEX": ({
        "referencia": ("numero de documento", "no de documento"),
        # El portal escribe la columna de las dos formas, con «de» y sin él
        # («Fecha Vencimiento»). Con una sola variante, un export legítimo no
        # casaba y las 3 127 facturas se perdían enteras.
        "fecha": ("fecha de vencimiento", "fecha vencimiento"),
        "importe": ("saldo",),
    }, ("referencia", "fecha", "importe")),
    "MGC": ({
        "referencia": ("referencia",),
        "fecha": ("vencimiento neto", "vencimiento"),
        "importe": ("importe en moneda local", "importe"),
    }, ("referencia", "fecha", "importe")),
    "TESORO": ({
        "referencia": ("factura",),
        "fecha": ("fecha vencimiento", "fecha de vencimiento"),
        "importe": ("monto",),
    }, ("referencia", "fecha", "importe")),
    # La nómina no tiene columna distintiva, pero «Fecha de pago» no aparece en
    # ningún reporte bancario.
    "NOMINA": ({
        "referencia": ("semana", "periodo"),
        "fecha": ("fecha de pago", "fecha pago"),
        "importe": ("importe", "monto", "total"),
    }, ("fecha",)),
}

# Rangos de CRÉDITOS que se copian tal cual. Abarcan TODA la hoja útil, rótulos
# incluidos, y no solo los importes: lo que el usuario entrega es la hoja entera y
# es esa la que debe quedar en el reporte. (No hay que confundirlos con los rangos
# que el derivador VACÍA para dejar el molde: aquellos se limitan a los importes,
# justo para que el esqueleto sobreviva.)
_RANGOS_CREDITOS = (
    (2, 5, "C", "D"),      # Fecha, Semana, Cobertura mes, Excedente cobertura
    (7, 139, "A", "F"),    # líneas de crédito por cuenta
    (1, 118, "J", "V"),    # matriz empresa x acreedor x día
)

_EXTENSIONES = (".xlsx", ".xls")


class ErrorInsumo(Exception):
    """El archivo no se pudo leer o no corresponde a ningún insumo."""


def _filas(ruta: str, hoja: str = None) -> list:
    """Filas de un archivo. Con `hoja`, las de esa pestaña en concreto."""
    ext = os.path.splitext(ruta)[1].lower()
    if ext == ".xls":
        return _filas_xls(ruta)
    if ext != ".xlsx":
        raise ErrorInsumo(
            "«{}»: los insumos de flujo se leen en Excel (.xlsx o .xls)".format(
                os.path.basename(ruta)))
    if hoja is None:
        return _filas_xlsx(ruta)
    import openpyxl
    libro = openpyxl.load_workbook(ruta, data_only=True)
    try:
        return [list(f) for f in libro[hoja].iter_rows(values_only=True)]
    finally:
        libro.close()


def es_libro_de_insumos(ruta: str) -> bool:
    """Si el archivo es el LIBRO de insumos, por los nombres de sus pestañas.

    Se mira solo el índice de hojas —que openpyxl entrega sin leer una sola
    fila— para no pagar la lectura completa del libro antes de saber qué es.
    Importa: el libro de insumos ronda los 2 MB y 32 000 filas, y cada apertura
    cuesta más de diez segundos.

    Basta con DOS secciones para reconocerlo: con una sola, un reporte bancario
    que por casualidad tuviera una pestaña llamada «NOMINA» se colaría."""
    hojas = {_clave(h) for h in hojas_de(ruta)}
    return len(hojas & set(SECCIONES)) >= 2


def hojas_de(ruta: str) -> list:
    """Nombres de las pestañas de un .xlsx (vacío si no se puede abrir)."""
    if os.path.splitext(ruta)[1].lower() != ".xlsx":
        return []
    import openpyxl
    try:
        libro = openpyxl.load_workbook(ruta, read_only=True)
    except Exception:  # noqa: BLE001 — no es un xlsx legible
        return []
    try:
        return list(libro.sheetnames)
    finally:
        libro.close()


def _fecha(valor):
    """Fecha comparable con las del calendario, o None.

    Los `SUMIF` de SALDOS comparan contra fechas sin hora, así que se trunca: un
    vencimiento con hora nunca casaría con el día del calendario."""
    if isinstance(valor, datetime.datetime):
        return datetime.datetime(valor.year, valor.month, valor.day)
    if isinstance(valor, datetime.date):
        return datetime.datetime(valor.year, valor.month, valor.day)
    if isinstance(valor, str):
        texto = valor.strip()
        for patron in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(texto[:10], patron)
            except ValueError:
                continue
    return None


def detectar(ruta: str, hoja: str = None) -> str:
    """Qué insumo es este archivo (o esta pestaña). None si no es ninguno."""
    if os.path.splitext(ruta)[1].lower() not in _EXTENSIONES:
        return None
    try:
        filas = _filas(ruta, hoja)
    except (ErrorInsumo, ErrorLector, KeyError):
        return None
    return _detectar_en(filas)


def _detectar_en(filas: list) -> str:
    """Igual que `detectar`, pero sobre filas ya extraídas.

    Existe para no reabrir el archivo: un libro con las seis secciones se cargaba
    una vez por pestaña para detectar y otra para leer — doce cargas de 32 000
    filas, sesenta segundos."""
    if _parece_creditos(filas):
        return "CREDITOS"
    mejor, mejor_n = None, 0
    for nombre, (alias, obligatorias) in _ALIAS.items():
        n, idx = _buscar_encabezado(filas, alias, obligatorias)
        # Gana el que reconozca MÁS columnas: 'Saldo' y 'Monto' aparecen en varios
        # reportes, y quedarse con el primero que casa elige mal.
        if n is not None and len(idx) > mejor_n:
            mejor, mejor_n = nombre, len(idx)
    return mejor


def _parece_creditos(filas: list) -> bool:
    """CRÉDITOS se reconoce por los encabezados de su bloque de líneas."""
    esperados = {"amor. cp", "semana cp", "semana lp"}
    for fila in filas[:12]:
        if not fila:
            continue
        vistos = {_norm(c) for c in fila if c is not None}
        if len(esperados & vistos) >= 2:
            return True
    return False


def leer_ledger(ruta: str, nombre: str, hoja: str = None) -> list:
    return _leer_ledger_en(_filas(ruta, hoja), nombre, os.path.basename(ruta))


def _leer_ledger_en(filas: list, nombre: str, origen: str = "") -> list:
    """Filas de un ledger de vencimientos, COMPLETAS.

    Cada registro lleva `fecha` e `importe` ya normalizados —son los que consumen
    los `SUMIF` de SALDOS— y además `celdas` con la fila entera tal como venía.

    Guardar la fila completa no es adorno: MGC exporta once columnas y Pemex
    trece, y con solo tres no hay forma de cotejar el reporte contra el sistema.
    Quien vea un total raro necesita el número de documento, la clase, la fecha de
    emisión y el texto de cabecera para ir a buscarlo.

    Se descartan las filas sin fecha o sin importe: no aportan a ningún `SUMIF`
    y solo harían más pesado el libro."""
    alias, obligatorias = _ALIAS[nombre]
    n, idx = _buscar_encabezado(filas, alias, obligatorias)
    if n is None:
        raise ErrorInsumo(
            "«{}» no trae las columnas de {} ({})".format(
                origen, nombre, ", ".join(obligatorias)))

    # El calendario de nómina trae el importe en la columna de junto a la fecha,
    # sin encabezado. Es el único insumo así, y es su forma de siempre.
    if "importe" not in idx and "fecha" in idx:
        idx["importe"] = idx["fecha"] + 1

    salida = []
    for fila in filas[n + 1:]:
        if not fila:
            continue
        fecha = _fecha(_valor(fila, idx.get("fecha")))
        if fecha is None:
            continue
        importe = _a_float(_valor(fila, idx.get("importe")))
        if importe is None:
            continue
        salida.append({
            "referencia": _valor(fila, idx.get("referencia")),
            "fecha": fecha,
            "importe": float(importe),
            # La fila tal cual. Se copia por POSICIÓN, igual que el pegado manual:
            # la pestaña del formato ES esa exportación pegada, así que sus
            # columnas coinciden una a una. Se recortan las vacías del final para
            # que releer el archivo no vaya inflando la fila con las columnas de
            # relleno que openpyxl reporta por el ancho de la hoja.
            "celdas": _sin_cola_vacia(fila),
        })
    return salida


def _con_datos(filas) -> bool:
    """Si la pestaña trae algo más que su fila de encabezados.

    La plantilla en blanco que se ofrece a descargar sale con los encabezados y
    nada más: esa no es un error, es el punto de partida."""
    llenas = [f for f in (filas or ()) if any(c is not None for c in f)]
    return len(llenas) > 1


def _sin_cola_vacia(fila: list) -> list:
    """La fila sin las celdas vacías del final."""
    fin = len(fila)
    while fin and fila[fin - 1] is None:
        fin -= 1
    return list(fila[:fin])


def _valor(fila: list, i):
    if i is None or i >= len(fila):
        return None
    return fila[i]


def leer_creditos(ruta: str, hoja: str = None) -> dict:
    """Los rangos de CRÉDITOS, tal cual, para copiarlos a las mismas celdas.

    Devuelve {"rangos": [{"fila_ini", "col_ini", "celdas": [[...], ...]}]}. Se
    lee con openpyxl directamente y no por `_filas` porque hacen falta las
    coordenadas, no solo los valores."""
    import openpyxl
    from openpyxl.utils import column_index_from_string

    try:
        libro = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as exc:  # noqa: BLE001 — se traduce a un error propio
        raise ErrorInsumo("No se pudo leer «{}»: {}".format(
            os.path.basename(ruta), exc)) from exc
    try:
        return _leer_creditos_hoja(libro[hoja] if hoja
                                   else _hoja_creditos(libro))
    finally:
        libro.close()


def _leer_creditos_hoja(hoja) -> dict:
    """Los rangos de CRÉDITOS de una hoja ya abierta."""
    from openpyxl.utils import column_index_from_string
    rangos = []
    for fila_ini, fila_fin, col_ini, col_fin in _RANGOS_CREDITOS:
        ci = column_index_from_string(col_ini)
        cf = column_index_from_string(col_fin)
        celdas = [[hoja.cell(f, c).value for c in range(ci, cf + 1)]
                  for f in range(fila_ini, fila_fin + 1)]
        rangos.append({"fila_ini": fila_ini, "col_ini": col_ini,
                       "celdas": celdas})
    return {"rangos": rangos}


def _hoja_creditos(libro):
    """La hoja de créditos: la que se llame así, o la primera que lo parezca."""
    import unicodedata
    for nombre in libro.sheetnames:
        plano = unicodedata.normalize("NFKD", nombre)
        plano = "".join(c for c in plano if not unicodedata.combining(c))
        if plano.upper().strip().startswith("CREDITO"):
            return libro[nombre]
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        filas = [[c.value for c in f]
                 for f in hoja.iter_rows(min_row=1, max_row=12)]
        if _parece_creditos(filas):
            return hoja
    raise ErrorInsumo(
        "«{}» no tiene una hoja de créditos reconocible".format(libro))


SECCIONES = ("CREDITOS", "MGC", "PEMEX", "TESORO", "NOMINA", "IMPUESTOS")

def leer(ruta: str, nombre: str = None) -> tuple:
    """Lee un archivo de insumos. Devuelve `(nombre, datos)`.

    Un archivo con VARIAS pestañas reconocibles devuelve
    `("COMBINADO", {seccion: datos})`. Antes solo se miraba la primera hoja, así
    que un libro con las seis secciones entregaba una y perdía cinco EN SILENCIO:
    el reporte salía con los paneles vacíos y nada lo decía.

    `nombre` fuerza el tipo; si no se pasa, se detecta por los encabezados."""
    if nombre:
        return nombre, (leer_creditos(ruta) if nombre == "CREDITOS"
                        else leer_ledger(ruta, nombre))

    hojas = hojas_de(ruta)
    if len(hojas) > 1:
        # UNA sola apertura para todo el libro. Antes se abría por pestaña y por
        # operación (detectar + leer), o sea doce veces un archivo de 32 000
        # filas: sesenta segundos que el usuario esperaba sin saber por qué.
        import openpyxl
        libro = openpyxl.load_workbook(ruta, data_only=True)
        try:
            por_hoja = {n: [list(f) for f in libro[n].iter_rows(values_only=True)]
                        for n in libro.sheetnames}
            combinado = {}

            def _tomar(hoja, tipo, exigir=False):
                if tipo is None or tipo in combinado:
                    return
                try:
                    combinado[tipo] = (
                        _leer_creditos_hoja(libro[hoja]) if tipo == "CREDITOS"
                        else _leer_ledger_en(por_hoja[hoja], tipo,
                                             os.path.basename(ruta)))
                except (ErrorInsumo, ErrorLector):
                    # Una pestaña que solo SE PARECE a un insumo y no se deja
                    # leer no invalida las demás. Pero una que se LLAMA como la
                    # sección y trae datos, sí: significa que el usuario la subió
                    # esperando que entrara. Callarlo dejaba su panel en cero sin
                    # decir nada, que es justo el fallo silencioso que este
                    # módulo existe para no repetir.
                    if exigir and _con_datos(por_hoja.get(hoja)):
                        raise

            # Dos pasadas, y el ORDEN importa. El nombre de la pestaña manda sobre
            # el olfateo de encabezados: si alguien sube el formato completo, su
            # hoja SALDOS lleva la misma cabecera 'Cta. / Amor. CP / Semana CP'
            # que CRÉDITOS —el panel de créditos vive dentro de SALDOS— y, por ir
            # antes en el libro, se quedaba con el puesto. El reporte salía con la
            # hoja de saldos metida en la pestaña de créditos y nada lo decía.
            pendientes = []
            for hoja in hojas:
                clave = _clave(hoja)
                if clave.startswith("SALDOS"):
                    continue   # SALDOS y SALDOS HORIZOTAL son salida, no insumo
                # IMPUESTOS está en SECCIONES pero no tiene lector: el formato
                # reserva la pestaña y SALDOS todavía no la consulta. Se deja
                # caer a la segunda pasada, que simplemente no la reconocerá.
                if clave == "CREDITOS" or clave in _ALIAS:
                    _tomar(hoja, clave, exigir=True)
                else:
                    pendientes.append(hoja)
            for hoja in pendientes:
                _tomar(hoja, _detectar_en(por_hoja.get(hoja) or []))
        finally:
            libro.close()
        if len(combinado) > 1:
            return "COMBINADO", combinado
        if combinado:
            tipo, datos = next(iter(combinado.items()))
            return tipo, datos

    nombre = detectar(ruta)
    if nombre is None:
        raise ErrorInsumo(
            "«{}» no parece ninguno de los insumos de flujo "
            "(créditos, Pemex, MGC, tesoro, nómina)".format(
                os.path.basename(ruta)))
    if nombre == "CREDITOS":
        return nombre, leer_creditos(ruta)
    return nombre, leer_ledger(ruta, nombre)


def escribir_plantilla(ruta: str, datos: dict = None) -> dict:
    """Escribe el libro de insumos: una pestaña por sección.

    Las pestañas NO se arman a mano: se copian del libro base, que es el formato
    de tesorería vaciado. Así los encabezados y el ancho de las columnas son los
    del formato real —MGC trae once columnas, Pemex trece— y no una versión
    recortada que alguien tendría que traducir al llenarla.

    Es a la vez la plantilla en blanco que se descarga y el archivo donde la app
    conserva lo capturado. Que sean el mismo evita la trampa de tener una copia
    interna distinta de la que ve el usuario.

    Devuelve {sección: filas escritas}."""
    import openpyxl
    from . import saldos_plantilla

    plantilla = saldos_plantilla.cargar()
    libro = openpyxl.load_workbook(plantilla.ruta_base)
    # Se queda solo con las pestañas de insumos. Ninguna tiene fórmulas, así que
    # quitar las demás no deja referencias rotas.
    quedan = {_clave(n): n for n in libro.sheetnames}
    for clave, nombre in list(quedan.items()):
        if clave not in SECCIONES:
            del libro[nombre]

    datos = datos or {}
    escritas = {}
    for seccion in SECCIONES:
        nombre = quedan.get(seccion)
        if nombre is None:
            continue
        hoja = libro[nombre]
        info = plantilla.ledgers.get(seccion) or {}
        contenido = datos.get(seccion)
        if info.get("modo") == "copia":
            escritas[seccion] = _copiar_rangos(hoja, contenido)
        else:
            escritas[seccion] = _volcar_ledger(hoja, info, contenido)

    libro.save(ruta)
    return escritas


def _clave(nombre: str) -> str:
    """'CRÉDITOS' -> 'CREDITOS', para casar el nombre de la hoja con la sección."""
    import unicodedata
    plano = unicodedata.normalize("NFKD", nombre)
    return "".join(c for c in plano if not unicodedata.combining(c)).upper().strip()


def _volcar_ledger(hoja, info: dict, filas) -> int:
    """Escribe un ledger: la fila COMPLETA, y encima los campos normalizados.

    Primero se vuelca `celdas` —todas las columnas que trajo la exportación— para
    que el archivo sirva para cotejar contra el sistema de origen. Después se
    reescriben la fecha y el importe con los valores ya interpretados, porque los
    `SUMIF` comparan contra fechas reales y una fecha en texto no casaría."""
    cols = info.get("cols") or {}
    if not filas:
        return 0
    from openpyxl.utils import column_index_from_string
    fila = info.get("fila_ini", 2)
    tope = info.get("fila_fin", fila + len(filas))
    escritas = 0
    for registro in filas:
        if fila > tope:
            break
        for i, valor in enumerate(registro.get("celdas") or (), start=1):
            if valor is not None:
                hoja.cell(fila, i).value = valor
        # Solo se reescriben fecha e importe, que son los que los `SUMIF`
        # necesitan interpretados. La referencia ya viene dentro de `celdas`, y
        # sobreescribirla movía la etiqueta de la columna A de NÓMINA.
        for rol in ("fecha", "importe"):
            letra, valor = cols.get(rol), registro.get(rol)
            if letra and valor is not None:
                hoja.cell(fila, column_index_from_string(letra)).value = valor
        fila += 1
        escritas += 1
    return escritas


def _copiar_rangos(hoja, datos) -> int:
    """Reescribe CRÉDITOS en sus coordenadas exactas.

    No lleva encabezados propios que inventar: los trae la pestaña copiada del
    formato. SALDOS la lee por posición, así que la geometría tiene que coincidir."""
    if not datos:
        return 0
    from openpyxl.utils import column_index_from_string
    total = 0
    for rango in datos.get("rangos", ()):
        col_ini = column_index_from_string(rango["col_ini"])
        for i, valores in enumerate(rango["celdas"]):
            fila = rango["fila_ini"] + i
            for j, valor in enumerate(valores):
                if valor is None:
                    continue
                hoja.cell(fila, col_ini + j).value = valor
                total += 1
    return total


def leer_varios(rutas) -> tuple:
    """Lee un lote. Devuelve `({nombre: datos}, [errores])`.

    Si dos archivos son del mismo insumo gana el último, que es lo que espera
    quien vuelve a subir un archivo para corregir el anterior."""
    insumos, errores = {}, []
    for ruta in rutas or ():
        try:
            nombre, datos = leer(ruta)
            insumos[nombre] = datos
        except (ErrorInsumo, ErrorLector) as exc:
            errores.append("{}: {}".format(os.path.basename(ruta), exc))
    return insumos, errores

"""Lectores de los reportes de saldos que emite cada portal bancario.

Cada banco entrega lo suyo en un formato distinto —csv, xlsx, xls antiguo, txt de
ancho fijo o pdf— y con encabezados propios. Este módulo los normaliza todos a una
misma lista de `LineaSaldo`, para que el resto del sistema no tenga que saber de
qué banco viene cada archivo.

El banco se detecta **por el contenido**, no por el nombre del archivo: cada lector
declara la firma de su reporte (los encabezados que espera) y se prueban en orden.
Así un archivo renombrado sigue funcionando, que es lo habitual cuando el usuario
descarga varios el mismo día. El nombre solo se usa como pista de desempate.

Principio de diseño: **si un reporte no se reconoce, se avisa; nunca se devuelven
saldos en cero**. Un cero silencioso en un reporte de tesorería es peor que un
error: se firma como bueno.

Formatos cubiertos y de dónde sale el dato (verificado contra descargas reales):

    BANORTE     csv        CUENTA · CLABE · SALDO DISPONIBLE
    SANTANDER   csv        Cuenta · Disponible
    BANAMEX     csv        Sucursal+Cuenta · Saldo · Moneda
    BANREGIO    xlsx       Cuenta · Empresa · Disponible
    MULTIVA     xlsx       Cuenta · Divisa · Saldo
    BAJIO       xlsx       encabezado en la fila 7; Cuentas de vista · Saldo Disponible
    HSBC        xlsx       Número de cuenta · Actual disponible   (ver _filas_xlsx)
    BANCOMER    xls        Cuenta · Divisa · Disponible           (requiere xlrd)
    SCOTIABANK  txt        ancho fijo de 134 caracteres
    MONEX       pdf        Contrato · Clabe · Total en pesos
    SABADELL    pdf        sección 'Cuentas' · Saldo disponible (las Líneas de
                           Crédito NO son saldos)

    INBURSA     xlsx       CUENTA · SALDO DISPONIBLE (la divisa va en la portada)
"""

from __future__ import annotations

import csv
import os
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

from .catalogo_bancos import banco_desde_clabe

try:
    import openpyxl
except ImportError:  # sin openpyxl no se pueden leer los reportes en xlsx
    openpyxl = None

try:
    import xlrd  # solo para el .xls antiguo de BANCOMER (BIFF/OLE2)
except ImportError:
    xlrd = None

try:
    import pymupdf
except ImportError:
    pymupdf = None


class ErrorLector(Exception):
    """No se pudo leer el reporte (formato no reconocido, archivo dañado…)."""


@dataclass
class LineaSaldo:
    """Un saldo tal como lo reporta el banco, ya normalizado.

    `cuenta` y `clabe` van en dígitos, sin recortar ceros a la izquierda: el casado
    posterior compara colas de dígitos y un cero perdido rompe la comparación.
    """

    banco: str            # nombre canónico del banco
    cuenta: str           # dígitos de la cuenta, tal como los reporta el portal
    clabe: str            # CLABE si el reporte la trae; "" si no
    titular: str          # nombre que aparece en el reporte del banco
    saldo: float
    moneda: str           # MXN | USD (u otra sigla, en mayúsculas)
    origen: str = ""      # archivo del que salió, para diagnóstico
    extra: dict = field(default_factory=dict)


# --------------------------------------------------------------- utilidades

def _norm(texto) -> str:
    """Minúsculas, sin acentos y con los espacios colapsados. Para comparar
    encabezados: los portales cambian tildes y espacios entre versiones."""
    plano = unicodedata.normalize("NFKD", str(texto or ""))
    plano = "".join(c for c in plano if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", plano).strip().lower()


def _digitos(valor) -> str:
    """Solo los dígitos de un valor. Quita el apóstrofo con que los portales
    fuerzan texto ('0502939411), guiones, espacios y máscaras."""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return re.sub(r"\D", "", str(valor or ""))


def _a_float(valor) -> float | None:
    """Monto a float. Acepta '$13,290,864.21', '1.234,56' no (los portales
    mexicanos usan punto decimal), negativos con signo o entre paréntesis, y
    devuelve None si no hay número (celda vacía, guion, texto)."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = str(valor).strip()
    negativo = txt.startswith("(") and txt.endswith(")")
    limpio = re.sub(r"[^\d.\-]", "", txt.replace(",", ""))
    if not limpio or limpio in ("-", ".", "-."):
        return None
    try:
        n = float(limpio)
    except ValueError:
        return None
    return -abs(n) if negativo else n


def _moneda(valor, defecto: str = "MXN") -> str:
    """Sigla de moneda normalizada. Los portales escriben MXP, MN, MXN, PESOS…"""
    s = re.sub(r"[^A-Z]", "", str(valor or "").upper())
    if s in ("MXP", "MN", "MXN", "PESOS", "PESO", "MEXICANO"):
        return "MXN"
    if s in ("USD", "DLLS", "DLL", "DOLARES", "DOLAR"):
        return "USD"
    return s or defecto


def _texto_plano(ruta: str, limite: int = 4096) -> str:
    """Primeros caracteres de un archivo de texto, probando codificaciones. Se usa
    para detectar el banco por sus encabezados."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(ruta, encoding=enc) as fh:
                return fh.read(limite)
        except (UnicodeDecodeError, LookupError):
            continue
        except OSError as exc:
            raise ErrorLector(f"No se pudo abrir «{os.path.basename(ruta)}»: {exc}")
    return ""


def _filas_csv(ruta: str) -> list[list[str]]:
    """Filas de un csv, con el delimitador deducido y tolerante a codificación."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(ruta, encoding=enc, newline="") as fh:
                muestra = fh.read(4096)
                fh.seek(0)
                try:
                    dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
                except csv.Error:
                    dialecto = csv.excel  # coma, el caso habitual
                return [f for f in csv.reader(fh, dialecto)
                        if any(x.strip() for x in f)]
        except UnicodeDecodeError:
            continue
    raise ErrorLector(f"No se pudo decodificar «{os.path.basename(ruta)}».")


def _filas_xlsx_crudo(ruta: str) -> list[list]:
    """Filas de un .xlsx leyendo el XML del zip con el parser de la stdlib.

    Existe porque los reportes de HSBC e INBURSA rompen openpyxl con
    `TypeError: expected <class 'openpyxl.styles.fills.Fill'>` — su hoja de estilos
    trae un relleno que openpyxl no acepta, y falla en TODAS las combinaciones de
    flags (read_only, data_only, rich_text). Los valores, en cambio, están intactos:
    basta con no mirar los estilos.

    Se usa ElementTree y NO expresiones regulares: el orden de los atributos de un
    elemento XML es arbitrario y de hecho cambia entre portales —HSBC escribe
    `<c r="A1" t="s">` e INBURSA `<c t="s" r="A1">`—, así que cualquier patrón que
    dé por sentado el orden lee bien un archivo y devuelve cero filas del otro.
    """
    try:
        with zipfile.ZipFile(ruta) as z:
            nombres = z.namelist()
            sst: list[str] = []
            if "xl/sharedStrings.xml" in nombres:
                raiz = ET.fromstring(z.read("xl/sharedStrings.xml"))
                for si in raiz.findall(f"{_NS}si"):
                    # Una cadena puede venir partida en varios <t> (texto con
                    # formato mezclado); se concatenan todos.
                    sst.append("".join(t.text or "" for t in si.iter(f"{_NS}t")))
            hojas = sorted(n for n in nombres
                           if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
            if not hojas:
                raise ErrorLector("El archivo no tiene hojas.")
            raiz = ET.fromstring(z.read(hojas[0]))
    except ErrorLector:
        raise
    except (zipfile.BadZipFile, ET.ParseError, KeyError, OSError) as exc:
        raise ErrorLector(
            f"No se pudo leer «{os.path.basename(ruta)}»: {exc}") from exc

    datos = raiz.find(f"{_NS}sheetData")
    if datos is None:
        return []
    filas: list[list] = []
    for fila in datos.findall(f"{_NS}row"):
        celdas: list = []
        for c in fila.findall(f"{_NS}c"):
            tipo = c.get("t")
            if tipo == "inlineStr":
                bloque = c.find(f"{_NS}is")
                valor = ("".join(t.text or "" for t in bloque.iter(f"{_NS}t"))
                         if bloque is not None else None)
            else:
                v = c.find(f"{_NS}v")
                valor = v.text if v is not None else None
                if tipo == "s" and valor is not None and valor.lstrip("-").isdigit():
                    i = int(valor)
                    valor = sst[i] if 0 <= i < len(sst) else ""
                elif tipo in (None, "n") and valor is not None:
                    valor = _numero_xml(valor)
            # Se respeta la COLUMNA declarada en la referencia: una celda vacía en
            # medio de la fila no debe correr las de la derecha.
            destino = _indice_columna(c.get("r") or "")
            if destino is None:
                celdas.append(valor)
                continue
            while len(celdas) < destino:
                celdas.append(None)
            celdas.append(valor)
        filas.append(celdas)
    return filas


def _indice_columna(ref: str) -> int | None:
    """Índice 0-based de la columna de una referencia tipo 'AB12'. None si no
    trae letras (algunas hojas omiten el atributo 'r')."""
    letras = re.match(r"([A-Z]+)", ref or "")
    if not letras:
        return None
    n = 0
    for ch in letras.group(1):
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _numero_xml(texto: str):
    """Valor numérico de una celda del XML: int si es entero, float si no.

    Se convierte para que este lector devuelva lo mismo que openpyxl; si dejara
    cadenas, un número de cuenta y un importe se comportarían distinto según por
    qué vía se leyó el archivo."""
    try:
        n = float(texto)
    except ValueError:
        return texto
    return int(n) if n.is_integer() else n


def _filas_xlsx(ruta: str) -> list[list]:
    """Filas de un .xlsx. Intenta openpyxl y cae al lector crudo si falla.

    No se usa `read_only=True`: con el reporte de BAJÍO devuelve solo la primera
    columna (su hoja declara mal las dimensiones), y ese modo no da ninguna ventaja
    con archivos de unas decenas de filas."""
    if openpyxl is not None:
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            try:
                ws = wb[wb.sheetnames[0]]
                return [list(f) for f in ws.iter_rows(values_only=True)]
            finally:
                wb.close()
        except Exception:  # noqa: BLE001 — se reintenta con el lector crudo
            pass
    return _filas_xlsx_crudo(ruta)


def _filas_xls(ruta: str) -> list[list]:
    """Filas de un .xls antiguo (BIFF/OLE2). Requiere xlrd."""
    if xlrd is None:
        raise ErrorLector(
            "Para leer los reportes .xls de BBVA hace falta la librería 'xlrd' "
            "(pip install xlrd).")
    try:
        wb = xlrd.open_workbook(ruta)
        ws = wb.sheet_by_index(0)
        return [[c.value for c in ws.row(r)] for r in range(ws.nrows)]
    except Exception as exc:  # noqa: BLE001 — se traduce a un error propio
        raise ErrorLector(
            f"No se pudo leer «{os.path.basename(ruta)}»: {exc}") from exc


def _texto_pdf(ruta: str) -> str:
    """Texto de un PDF. Los reportes bancarios traen capa de texto, así que no se
    invoca OCR: si algún día llega uno escaneado, `core.ocr.extraer_texto` es la
    vía, pero cuesta segundos por página y aquí no hace falta."""
    if pymupdf is None:
        raise ErrorLector("Falta PyMuPDF para leer los reportes en PDF.")
    try:
        with pymupdf.open(ruta) as doc:
            return "\n".join(p.get_text() for p in doc)
    except Exception as exc:  # noqa: BLE001 — se traduce a un error propio
        raise ErrorLector(
            f"No se pudo leer «{os.path.basename(ruta)}»: {exc}") from exc


def _indice_encabezado(fila: list, alias: dict[str, tuple]) -> dict[str, int]:
    """Mapea nombre lógico -> índice de columna, buscando por encabezado.

    `alias` es {nombre_logico: (variantes aceptadas, ya normalizadas)}. Se compara
    por igualdad y, si no, por 'empieza con', porque algunos portales le pegan
    unidades o notas al encabezado."""
    idx: dict[str, int] = {}
    normalizadas = [_norm(c) for c in fila]
    for logico, variantes in alias.items():
        for i, celda in enumerate(normalizadas):
            if celda in variantes or any(celda.startswith(v) for v in variantes):
                idx[logico] = i
                break
    return idx


def _buscar_encabezado(filas: list[list], alias: dict[str, tuple],
                       obligatorias: tuple, limite: int = 25):
    """Encuentra la fila de encabezados y devuelve `(nº de fila, índice)`.

    No se asume la fila 1: el reporte de BAJÍO trae seis filas de portada y sus
    encabezados empiezan en la 7. Se escanea hasta `limite` filas buscando la
    primera que contenga todas las columnas obligatorias."""
    for n, fila in enumerate(filas[:limite]):
        if not fila:
            continue
        idx = _indice_encabezado(fila, alias)
        if all(o in idx for o in obligatorias):
            return n, idx
    return None, {}


# ------------------------------------------------------------------ lectores
# Cada lector devuelve list[LineaSaldo] o lanza ErrorLector. La detección del
# banco va aparte, en _detectar(): así un lector se puede forzar a mano.

_ALIAS_BANORTE = {
    "cuenta": ("cuenta",),
    "titular": ("titular / personalizacion", "titular"),
    "moneda": ("moneda",),
    "clabe": ("clabe",),
    "saldo": ("saldo disponible",),
}


def leer_banorte(ruta: str) -> list[LineaSaldo]:
    filas = _filas_csv(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_BANORTE, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de BANORTE.")
    out = []
    for fila in filas[n + 1:]:
        cuenta = _digitos(fila[idx["cuenta"]]) if idx["cuenta"] < len(fila) else ""
        if not cuenta:
            continue
        clabe = _digitos(fila[idx["clabe"]]) if "clabe" in idx else ""
        saldo = _a_float(fila[idx["saldo"]])
        out.append(LineaSaldo(
            banco="Banorte", cuenta=cuenta, clabe=clabe,
            titular=str(fila[idx["titular"]] if "titular" in idx else "").strip(),
            saldo=saldo or 0.0,
            moneda=_moneda(fila[idx["moneda"]] if "moneda" in idx else "")))
    return out


_ALIAS_SANTANDER = {
    "cuenta": ("cuenta",),
    "titular": ("descripcion",),
    "saldo": ("disponible",),
}


def leer_santander(ruta: str) -> list[LineaSaldo]:
    filas = _filas_csv(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_SANTANDER, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de SANTANDER.")
    out = []
    for fila in filas[n + 1:]:
        cuenta = _digitos(fila[idx["cuenta"]]) if idx["cuenta"] < len(fila) else ""
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="Santander", cuenta=cuenta, clabe="",
            titular=str(fila[idx["titular"]] if "titular" in idx else "").strip(),
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda="MXN"))  # el consolidado de cheques no trae columna de moneda
    return out


_ALIAS_BANAMEX = {
    "sucursal": ("sucursal",),
    "cuenta": ("cuenta",),
    "saldo": ("saldo",),
    "moneda": ("moneda",),
    "error": ("mensaje de error",),
}


def leer_banamex(ruta: str) -> list[LineaSaldo]:
    filas = _filas_csv(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_BANAMEX, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de BANAMEX.")
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        if not cuenta:
            continue
        # El número completo es sucursal + cuenta: así es como está en el catálogo
        # (p. ej. sucursal 394 + cuenta 7680454 -> 3947680454).
        sucursal = _digitos(fila[idx["sucursal"]]) if "sucursal" in idx else ""
        out.append(LineaSaldo(
            banco="Banamex", cuenta=sucursal + cuenta, clabe="", titular="",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda=_moneda(fila[idx["moneda"]] if "moneda" in idx else ""),
            extra={"cuenta_corta": cuenta, "sucursal": sucursal,
                   "aviso": str(fila[idx["error"]]).strip()
                   if "error" in idx and idx["error"] < len(fila) else ""}))
    return out


_ALIAS_BANREGIO = {
    "alias": ("alias",),
    "cuenta": ("cuenta",),
    "titular": ("empresa",),
    "saldo": ("disponible",),
}


def leer_banregio(ruta: str) -> list[LineaSaldo]:
    filas = _filas_xlsx(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_BANREGIO, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de BANREGIO.")
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="Banregio", cuenta=cuenta, clabe="",
            titular=str(fila[idx["titular"]] or "").strip() if "titular" in idx else "",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            # El reporte no trae divisa; el catálogo distingue las cuentas en
            # dólares, así que la moneda se resuelve al casar, no aquí.
            moneda="MXN",
            extra={"alias": str(fila[idx["alias"]] or "").strip()
                   if "alias" in idx and idx["alias"] < len(fila) else ""}))
    return out


_ALIAS_MULTIVA = {
    "titular": ("empresa",),
    "cuenta": ("cuenta",),
    "alias": ("alias",),
    "moneda": ("divisa",),
    "saldo": ("saldo",),
}


def leer_multiva(ruta: str) -> list[LineaSaldo]:
    filas = _filas_xlsx(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_MULTIVA, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de MULTIVA.")
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="Multiva Banco", cuenta=cuenta, clabe="",
            titular=str(fila[idx["titular"]] or "").strip() if "titular" in idx else "",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda=_moneda(fila[idx["moneda"]] if "moneda" in idx else "")))
    return out


_ALIAS_BAJIO = {
    "producto": ("tipo de producto",),
    "cuenta": ("cuentas de vista",),
    "titular": ("nombre del cliente",),
    "moneda": ("divisa",),
    "saldo": ("saldo disponible",),
}


def leer_bajio(ruta: str) -> list[LineaSaldo]:
    filas = _filas_xlsx(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_BAJIO, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de BAJÍO.")
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        # El mismo reporte mezcla cuentas con líneas de crédito y tarjetas; solo las
        # cuentas son saldo disponible. Sumar una línea de crédito inflaría el
        # reporte con dinero que no existe.
        producto = _norm(fila[idx["producto"]]) if "producto" in idx else "cuenta"
        if producto and not producto.startswith("cuenta"):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="Banco del Bajío", cuenta=cuenta, clabe="",
            titular=str(fila[idx["titular"]] or "").strip() if "titular" in idx else "",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda=_moneda(fila[idx["moneda"]] if "moneda" in idx else "")))
    return out


_ALIAS_HSBC = {
    "moneda": ("moneda",),
    "cuenta": ("numero de cuenta",),
    "titular": ("nombre de cuenta",),
    "saldo": ("actual disponible",),
}


def leer_hsbc(ruta: str) -> list[LineaSaldo]:
    filas = _filas_xlsx(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_HSBC, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de HSBC.")
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="HSBC", cuenta=cuenta, clabe="",
            titular=str(fila[idx["titular"]] or "").strip() if "titular" in idx else "",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda=_moneda(fila[idx["moneda"]] if "moneda" in idx else "")))
    return out


_ALIAS_INBURSA = {
    "cuenta": ("cuenta",),
    "titular": ("empresa",),
    "producto": ("producto",),
    "saldo": ("saldo disponible",),
}


def leer_inbursa(ruta: str) -> list[LineaSaldo]:
    """Inbursa entrega un consolidado por divisa: la moneda no está en una columna
    sino en una línea de portada ('Divisa: PESOS'), y la última fila es el total
    (sin número de cuenta, así que se descarta sola)."""
    filas = _filas_xlsx(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_INBURSA, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de INBURSA.")
    portada = _norm(" ".join(str(v) for f in filas[:n] for v in f if v))
    moneda = "USD" if ("dolar" in portada or "usd" in portada) else "MXN"
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="Inbursa", cuenta=cuenta, clabe="",
            titular=str(fila[idx["titular"]] or "").strip() if "titular" in idx else "",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda=moneda,
            extra={"producto": str(fila[idx["producto"]] or "").strip()
                   if "producto" in idx and idx["producto"] < len(fila) else ""}))
    return out


_ALIAS_BANCOMER = {
    "cuenta": ("cuenta",),
    "alias": ("alias",),
    "moneda": ("divisa",),
    "saldo": ("disponible",),
}


def leer_bancomer(ruta: str) -> list[LineaSaldo]:
    filas = _filas_xls(ruta) if ruta.lower().endswith(".xls") else _filas_xlsx(ruta)
    n, idx = _buscar_encabezado(filas, _ALIAS_BANCOMER, ("cuenta", "saldo"))
    if n is None:
        raise ErrorLector("No se encontraron los encabezados de BBVA.")
    out = []
    for fila in filas[n + 1:]:
        if idx["cuenta"] >= len(fila):
            continue
        cuenta = _digitos(fila[idx["cuenta"]])
        # La última fila suele ser 'Totales'; sin dígitos de cuenta se descarta sola.
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="BBVA México", cuenta=cuenta, clabe="",
            titular=str(fila[idx["alias"]] or "").strip() if "alias" in idx else "",
            saldo=_a_float(fila[idx["saldo"]]) or 0.0,
            moneda=_moneda(fila[idx["moneda"]] if "moneda" in idx else "")))
    return out


# SCOTIABANK exporta ancho fijo de 134 caracteres, sin encabezado:
#   0-2 producto (CHQ) · 3-5 moneda · 6-14 plaza y coma · 16-35 cuenta (20 dígitos)
#   36-85 nombre · 86-102 saldo · 103+ país y estatus
_SCOTIA = re.compile(
    r"^(?P<producto>[A-Z]{3})(?P<moneda>[A-Z]{3})(?P<plaza>[A-Z ]+),\s*"
    r"(?P<cuenta>\d{12,24})(?P<titular>.{1,60}?)\s*"
    r"(?P<saldo>\d{6,}\.\d{2})(?P<pais>[A-Za-z]+?)(?P<estatus>Activa|No existe.*)?\s*$")


def leer_scotiabank(ruta: str) -> list[LineaSaldo]:
    texto = _texto_plano(ruta, limite=1_000_000)
    out = []
    for linea in texto.splitlines():
        if not linea.strip():
            continue
        m = _SCOTIA.match(linea.rstrip())
        if not m:
            continue
        cuenta = m.group("cuenta").lstrip("0")
        if not cuenta:
            continue
        out.append(LineaSaldo(
            banco="Scotiabank", cuenta=cuenta, clabe="",
            titular=m.group("titular").strip(),
            saldo=_a_float(m.group("saldo")) or 0.0,
            moneda=_moneda(m.group("moneda")),
            extra={"estatus": (m.group("estatus") or "").strip()}))
    if not out:
        raise ErrorLector("Ninguna línea del archivo tiene el formato de SCOTIABANK.")
    return out


_MONEX_CLIENTE = re.compile(r"Cliente\s*:\s*(.+)")
_MONEX_CONTRATO = re.compile(r"Contrato\s*:\s*(\d+)")
_MONEX_CLABE = re.compile(r"Clabe\s*:\s*(\d{18})")
_MONEX_TOTAL = re.compile(r"Total en pesos\s*\n\s*([\d,.]+)")


def leer_monex(ruta: str) -> list[LineaSaldo]:
    """Monex emite un PDF por contrato, no un consolidado.

    Se toma 'Total en pesos', que es el equivalente a la columna 'VALUACIÓN TOTAL'
    que el formato usa hoy (SALDOS!I22 apunta a MONEX!F3)."""
    texto = _texto_pdf(ruta)
    contrato = _MONEX_CONTRATO.search(texto)
    clabe = _MONEX_CLABE.search(texto)
    total = _MONEX_TOTAL.search(texto)
    if not (contrato or clabe):
        raise ErrorLector("El PDF no parece un reporte de saldos de MONEX.")
    if total is None:
        raise ErrorLector(
            f"No se encontró el 'Total en pesos' en "
            f"«{os.path.basename(ruta)}».")
    cliente = _MONEX_CLIENTE.search(texto)
    return [LineaSaldo(
        banco="Banco Monex",
        cuenta=contrato.group(1) if contrato else "",
        clabe=clabe.group(1) if clabe else "",
        titular=cliente.group(1).strip() if cliente else "",
        saldo=_a_float(total.group(1)) or 0.0,
        moneda="MXN")]


def leer_sabadell(ruta: str) -> list[LineaSaldo]:
    """Sabadell entrega la 'Posición Global': cuentas primero, líneas de crédito
    después. Solo la primera sección son saldos — una línea de crédito es dinero
    disponible para pedir prestado, no dinero en la cuenta.

    Se toma **Saldo disponible**, confirmado con tesorería. Ojo si alguien compara
    con el formato viejo: su fórmula apuntaba a SABADELL!E9, que por los
    encabezados de esa hoja parecería ser 'Saldo por aplicar'. Es un espejismo del
    pegado manual —la columna quedó recorrida—, no el criterio real."""
    texto = _texto_pdf(ruta)
    if "Posición Global" not in texto and "Saldo disponible" not in texto:
        raise ErrorLector("El PDF no parece un reporte de SABADELL.")
    lineas = [x.strip() for x in texto.splitlines()]
    # La sección de cuentas termina donde empiezan las líneas de crédito.
    try:
        fin = next(i for i, x in enumerate(lineas) if "Líneas de Crédito" in x)
    except StopIteration:
        fin = len(lineas)
    out = []
    for i, x in enumerate(lineas[:fin]):
        # Una cuenta es una corrida de dígitos seguida de la moneda y dos importes.
        if not re.fullmatch(r"\d{6,}", x):
            continue
        if i + 3 >= fin:
            continue
        moneda, por_aplicar, disponible = lineas[i + 1], lineas[i + 2], lineas[i + 3]
        if not re.fullmatch(r"[A-Z]{3}", moneda):
            continue
        saldo = _a_float(disponible)
        if saldo is None:
            continue
        out.append(LineaSaldo(
            banco="Banco Sabadell", cuenta=x.lstrip("0") or x, clabe="",
            titular=lineas[i - 1].strip() if i else "",
            saldo=saldo, moneda=_moneda(moneda),
            extra={"saldo_por_aplicar": _a_float(por_aplicar)}))
    if not out:
        raise ErrorLector(
            f"No se encontraron cuentas en «{os.path.basename(ruta)}».")
    return out


# --------------------------------------------------------------- detección
# (nombre, extensiones, marcas que deben aparecer en el texto, lector)
# El orden importa: se prueba de la marca más específica a la más genérica.
_LECTORES = (
    ("BANORTE", (".csv",), ("titular / personalizacion", "saldo disponible"),
     leer_banorte),
    ("SANTANDER", (".csv",), ("descripcion", "disponible", "sbc"), leer_santander),
    ("BANAMEX", (".csv",), ("tipo de cuenta", "sucursal", "mensaje de error"),
     leer_banamex),
    ("HSBC", (".xlsx",), ("numero de cuenta", "actual disponible"), leer_hsbc),
    ("BAJIO", (".xlsx",), ("cuentas de vista", "saldo disponible"), leer_bajio),
    ("INBURSA", (".xlsx",), ("saldo consolidado", "salvo buen cobro"),
     leer_inbursa),
    ("BANREGIO", (".xlsx",), ("alias", "empresa", "disponible", "en transito"),
     leer_banregio),
    ("MULTIVA", (".xlsx",), ("empresa", "cuenta", "alias", "divisa", "saldo"),
     leer_multiva),
    ("BANCOMER", (".xls", ".xlsx"), ("cuenta", "alias", "divisa", "disponible"),
     leer_bancomer),
    ("SCOTIABANK", (".txt",), ("chq",), leer_scotiabank),
    ("MONEX", (".pdf",), ("contrato", "clabe"), leer_monex),
    ("SABADELL", (".pdf",), ("posicion global", "saldo disponible"), leer_sabadell),
)


def _huella(ruta: str) -> str:
    """Texto representativo del archivo, normalizado, para detectar el banco."""
    ext = os.path.splitext(ruta)[1].lower()
    try:
        if ext in (".csv", ".txt"):
            return _norm(_texto_plano(ruta))
        if ext == ".pdf":
            return _norm(_texto_pdf(ruta)[:4000])
        if ext == ".xls":
            filas = _filas_xls(ruta)[:12]
        else:
            filas = _filas_xlsx(ruta)[:15]
        return _norm(" ".join(str(v) for f in filas for v in f if v is not None))
    except ErrorLector:
        raise
    except Exception:  # noqa: BLE001 — un archivo ilegible se reporta al detectar
        return ""


def detectar(ruta: str) -> str | None:
    """Nombre del banco cuyo reporte parece ser `ruta`, o None.

    Se decide por el CONTENIDO. El nombre del archivo solo desempata cuando dos
    firmas encajan (p. ej. BANREGIO y MULTIVA comparten 'empresa/cuenta/alias')."""
    ext = os.path.splitext(ruta)[1].lower()
    huella = _huella(ruta)
    if not huella:
        return None
    candidatos = [(n, m) for n, exts, m, _ in _LECTORES
                  if ext in exts and all(x in huella for x in m)]
    if not candidatos:
        return None
    if len(candidatos) > 1:
        nombre_archivo = _norm(os.path.basename(ruta))
        for nombre, _ in candidatos:
            if _norm(nombre) in nombre_archivo:
                return nombre
        # Sin pista en el nombre, gana la firma más específica (más marcas).
        candidatos.sort(key=lambda c: -len(c[1]))
    return candidatos[0][0]


_POR_NOMBRE = {n: f for n, _, _, f in _LECTORES}


def leer(ruta: str, banco: str | None = None) -> tuple[list[LineaSaldo], str]:
    """Lee un reporte y devuelve `(líneas, banco detectado)`.

    `banco` fuerza el lector cuando la detección falla (p. ej. un portal que cambió
    sus encabezados). Lanza ErrorLector si no se reconoce o si el lector falla."""
    if not os.path.exists(ruta):
        raise ErrorLector(f"No se encontró el archivo «{ruta}».")
    nombre = banco or detectar(ruta)
    if nombre is None:
        raise ErrorLector(
            f"No se reconoce de qué banco es «{os.path.basename(ruta)}». "
            "Puede ser un formato nuevo del portal.")
    lector = _POR_NOMBRE.get(nombre)
    if lector is None:
        raise ErrorLector(f"No hay lector para «{nombre}».")
    lineas = lector(ruta)
    origen = os.path.basename(ruta)
    for x in lineas:
        x.origen = origen
        # Si el reporte trae CLABE, el banco sale de ella: es el dato duro. El
        # nombre del lector es solo la firma del formato.
        if x.clabe:
            canonico = banco_desde_clabe(x.clabe)
            if canonico:
                x.banco = canonico
    return lineas, nombre


def leer_varios(rutas, progreso=None) -> tuple[list[LineaSaldo], list[str]]:
    """Lee varios reportes. Devuelve `(líneas, errores)`.

    Un archivo que falle NO tumba el lote: se reporta y se sigue con los demás, que
    es lo que hace falta cuando se cargan doce reportes de golpe. `progreso(hechos,
    total)` se llama tras cada archivo.
    """
    lineas: list[LineaSaldo] = []
    errores: list[str] = []
    total = len(rutas)
    for i, ruta in enumerate(rutas, 1):
        try:
            nuevas, _ = leer(ruta)
            lineas.extend(nuevas)
        except ErrorLector as exc:
            errores.append(str(exc))
        except Exception as exc:  # noqa: BLE001 — un archivo raro no aborta el lote
            errores.append(f"«{os.path.basename(ruta)}»: {exc}")
        if progreso is not None:
            progreso(i, total)
    return lineas, errores
